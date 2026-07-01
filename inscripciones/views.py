import json
import re
from copy import copy

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import requests
from django.core.paginator import Paginator
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from django.db import IntegrityError
from django.core.exceptions import ValidationError
from .models import Equipo, Cancha
from .forms import RegistroEquipoForm, EditarEquipoForm, RegistroJugadorForm, CargaMasivaJugadoresForm, EditarJugadorEntrenadorForm, EditarPerfilJugadorForm, CanchaForm, CargaMasivaCanchasForm, POSICIONES, PIES
import csv
import io
from datetime import date, datetime, timedelta
import time
import openpyxl
from openpyxl.utils.datetime import from_excel
from django.contrib.auth.hashers import make_password
from accounts.models import Usuario, Jugador
from django.contrib.auth import update_session_auth_hash
from django.utils import timezone
from .utils import validar_edad_categoria, _enviar_credenciales_jugador, geodificar_direccion, enviar_credenciales_jugadores_lote
from .constants import MENSAJE_ELIMINACION_PROGRAMADA
from .seleccion_equipo import equipo_activo, equipos_del_entrenador, seleccionar_equipo

MOTIVOS_RECHAZO_EQUIPO = [
    ('DOCUMENTOS', 'Documentacion incompleta'),
    ('DATOS', 'Datos inconsistentes'),
    ('CATEGORIA', 'Categoria no corresponde'),
    ('DUPLICADO', 'Equipo duplicado'),
    ('CONTACTO', 'No fue posible contactar al entrenador'),
    ('OTRO', 'Otro motivo'),
]

MOTIVOS_ELIMINACION_EQUIPO = [
    ('SOLICITUD', 'Solicitud del entrenador'),
    ('INACTIVIDAD', 'Inactividad prolongada'),
    ('INCUMPLIMIENTO', 'Incumplimiento de normas'),
    ('DATOS', 'Datos falsos o inconsistentes'),
    ('DISCIPLINA', 'Problemas disciplinarios'),
    ('OTRO', 'Otro motivo'),
]


def _construir_motivo(opcion, detalle, opciones):
    opcion = (opcion or '').strip()
    detalle = (detalle or '').strip()
    etiquetas = dict(opciones)
    if opcion == 'OTRO':
        return detalle
    etiqueta = etiquetas.get(opcion, '').strip()
    if etiqueta and detalle:
        return f'{etiqueta}: {detalle}'
    return etiqueta or detalle


def _pdf_escape(value):
    return str(value).replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _generar_pdf_equipos(equipos, filtros=None):
    filtros = filtros or {}
    filtros_aplicados = ', '.join(
        f'{clave}: {valor}' for clave, valor in filtros.items() if valor
    ) or 'Sin filtros'
    lines = [
        'Reporte de equipos F.A.S',
        f'Fecha: {date.today().strftime("%d/%m/%Y")}',
        f'Filtros: {filtros_aplicados}',
        f'Total equipos: {len(equipos)}',
        '',
        'Equipo | Estado | Categoria | Localidad | Entrenador | Correo',
    ]
    for equipo in equipos[:44]:
        lines.append(
            f'{equipo.nombre} | {equipo.estado_display} | {equipo.categoria_display} | '
            f'{equipo.localidad} | {equipo.entrenador.nombres} {equipo.entrenador.apellidos} | {equipo.entrenador.email}'
        )

    stream = ['BT', '/F1 10 Tf', '40 790 Td', '13 TL']
    for index, line in enumerate(lines):
        if index:
            stream.append('T*')
        stream.append(f'({_pdf_escape(line)}) Tj')
    stream.append('ET')
    content = '\n'.join(stream).encode('latin-1', errors='replace')
    objects = [
        b'<< /Type /Catalog /Pages 2 0 R >>',
        b'<< /Type /Pages /Kids [3 0 R] /Count 1 >>',
        b'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>',
        b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
        b'<< /Length ' + str(len(content)).encode('ascii') + b' >>\nstream\n' + content + b'\nendstream',
    ]
    output = io.BytesIO()
    output.write(b'%PDF-1.4\n')
    offsets = [0]
    for number, obj in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f'{number} 0 obj\n'.encode('ascii'))
        output.write(obj)
        output.write(b'\nendobj\n')
    xref = output.tell()
    output.write(f'xref\n0 {len(objects) + 1}\n'.encode('ascii'))
    output.write(b'0000000000 65535 f \n')
    for offset in offsets[1:]:
        output.write(f'{offset:010d} 00000 n \n'.encode('ascii'))
    output.write(f'trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF'.encode('ascii'))
    return output.getvalue()

def _get_equipo_entrenador(request):
    """Compatibilidad interna: retorna el equipo activo del entrenador."""
    return equipo_activo(request)


def _agendar_eliminacion_equipo(request, equipo, motivo):
    ahora = timezone.now()

    if equipo.eliminar_programada_para and equipo.eliminar_programada_para > ahora:
        messages.info(request, MENSAJE_ELIMINACION_PROGRAMADA)
        return False

    motivo_limpio = (motivo or '').strip()
    if equipo.estado == Equipo.Estado.APROBADO and not motivo_limpio:
        messages.error(request, 'Debes indicar el motivo de la eliminación.')
        return False

    if not motivo_limpio:
        if equipo.estado == Equipo.Estado.RECHAZADO and equipo.motivo_rechazo:
            motivo_limpio = equipo.motivo_rechazo
        else:
            motivo_limpio = 'Eliminacion programada'

    equipo.eliminar_programada_para = ahora + timedelta(days=3)
    equipo.motivo_eliminacion = motivo_limpio
    equipo.save()
    messages.success(request, 'Eliminación programada para 3 días.')
    return True


@login_required
def registrar_equipo(request):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    form = RegistroEquipoForm(
        request.POST or None,
        request.FILES or None,
        entrenador=request.user.entrenador,
    )

    if request.method == 'POST':
        if form.is_valid():
            data = form.cleaned_data
            try:
                equipo = Equipo()
                equipo.nombre         = data['nombre']
                equipo.descripcion    = data.get('descripcion', '')
                equipo.anio_fundacion = data['anio_fundacion']
                equipo.categoria      = data['categoria']
                equipo.localidad      = data['localidad']
                equipo.barrio         = data['barrio']
                equipo.entrenador     = request.user.entrenador
                equipo.estado         = Equipo.Estado.ESPERA
                if data.get('logo'):
                    equipo.logo = data['logo']
                equipo.full_clean()
                equipo.save()
                seleccionar_equipo(request, equipo.pk)
                messages.success(request, 'Equipo registrado correctamente. En espera de aprobación.')
                return redirect('inscripciones:mi_equipo')

            except ValidationError as e:
                if hasattr(e, 'message_dict'):
                    for field, errors in e.message_dict.items():
                        for error in errors:
                            if 'already exists' in error:
                                continue
                            if field == '__all__':
                                form.add_error(None, error)
                            elif field == '_categoria':
                                form.add_error('categoria', error)
                            else:
                                form.add_error(field, error)
                else:
                    for error in e.messages:
                        if 'already exists' not in error:
                            form.add_error(None, error)
            except ValueError as e:
                form.add_error(None, str(e))
            except IntegrityError as error:
                detalle = str(error).lower()
                if 'entrenador' in detalle or 'nombre' in detalle or 'duplicate' in detalle:
                    form.add_error(
                        None,
                        'La base de datos todavía conserva las restricciones antiguas de equipo. '
                        'Ejecuta las migraciones pendientes y vuelve a intentarlo.'
                    )
                else:
                    form.add_error(None, 'No fue posible registrar el equipo por un conflicto de datos.')

    return render(request, 'inscripciones/registrar_equipo.html', {'form': form})


@login_required
def mi_equipo(request):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')
    equipo = _get_equipo_entrenador(request)
    equipos = equipos_del_entrenador(request.user)
    puede_editar = True
    if equipo:
        ahora = timezone.now()
        if equipo.eliminar_programada_para and ahora < equipo.eliminar_programada_para:
            puede_editar = False
        if equipo.estado == Equipo.Estado.RECHAZADO:
            if equipo.bloqueado_hasta and ahora < equipo.bloqueado_hasta:
                puede_editar = False
    return render(request, 'inscripciones/mi_equipo.html', {
        'equipo': equipo,
        'equipos': equipos,
        'puede_editar': puede_editar,
    })


@login_required
def seleccionar_equipo_activo(request, equipo_id):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')
    equipo = seleccionar_equipo(request, equipo_id)
    messages.success(request, f'Ahora administras el equipo {equipo.nombre}.')
    destino = request.POST.get('next') or request.GET.get('next') or 'dashboard_entrenador'
    destinos_permitidos = {
        'dashboard_entrenador',
        'inscripciones:mi_equipo',
        'inscripciones:lista_jugadores',
        'lista_entrenamientos',
        'torneos:entrenador_lista_torneos',
        'torneos:entrenador_reporte_jugadores',
    }
    if destino not in destinos_permitidos:
        destino = 'dashboard_entrenador'
    return redirect(destino)


@login_required
def editar_equipo(request, equipo_id):
    equipo = get_object_or_404(Equipo, id=equipo_id)

    if request.user.rol != 'ENTRENADOR' or equipo.entrenador != request.user.entrenador:
        messages.error(request, 'No tienes permiso para editar este equipo.')
        return redirect('inscripciones:mi_equipo')

    ahora = timezone.now()
    if equipo.eliminar_programada_para and ahora < equipo.eliminar_programada_para:
        messages.error(request, 'Este equipo tiene una eliminación programada.')
        return redirect('inscripciones:mi_equipo')

    if equipo.estado == Equipo.Estado.RECHAZADO:
        if equipo.bloqueado_hasta and ahora < equipo.bloqueado_hasta:
            messages.error(request, 'Este equipo está rechazado y no puede editarse todavía.')
            return redirect('inscripciones:mi_equipo')

    form = EditarEquipoForm(
        request.POST or None,
        request.FILES or None,
        equipo_pk=equipo.id,
        entrenador=request.user.entrenador,
        initial={
            'nombre':         equipo.nombre,
            'descripcion':    equipo.descripcion,
            'anio_fundacion': equipo.anio_fundacion,
            'categoria':      equipo.categoria,
            'localidad':      equipo.localidad,
            'barrio':         equipo.barrio,
        }
    )

    if request.method == 'POST':
        if form.is_valid():
            data = form.cleaned_data
            try:
                equipo.nombre         = data['nombre']
                equipo.descripcion    = data.get('descripcion', '')
                equipo.anio_fundacion = data['anio_fundacion']
                equipo.categoria      = data['categoria']
                equipo.localidad      = data['localidad']
                equipo.barrio         = data['barrio']
                if data.get('logo'):
                    equipo.logo = data['logo']
                equipo.estado = Equipo.Estado.ESPERA
                equipo.motivo_rechazo = None
                equipo.fecha_rechazo = None
                equipo.bloqueado_hasta = None
                equipo.eliminar_programada_para = None
                equipo.full_clean()
                equipo.save()
                messages.success(request, 'Equipo actualizado. En espera de aprobación.')
                return redirect('inscripciones:mi_equipo')

            except ValidationError as e:
                if hasattr(e, 'message_dict'):
                    for field, errors in e.message_dict.items():
                        for error in errors:
                            if 'already exists' in error:
                                continue
                            if field == '__all__':
                                form.add_error(None, error)
                            elif field == '_categoria':
                                form.add_error('categoria', error)
                            else:
                                form.add_error(field, error)
                else:
                    for error in e.messages:
                        if 'already exists' not in error:
                            form.add_error(None, error)
            except ValueError as e:
                form.add_error(None, str(e))
            except IntegrityError:
                form.add_error(None, 'No fue posible actualizar el equipo por un conflicto de datos.')

    return render(request, 'inscripciones/editar_equipo.html', {
        'form':   form,
        'equipo': equipo,
    })

def api_localidades(request):
    localidades = [
        'Antonio Nariño', 'Barrios Unidos', 'Bosa', 'Chapinero',
        'Ciudad Bolívar', 'Engativá', 'Fontibón', 'Kennedy',
        'La Candelaria', 'Los Mártires', 'Puente Aranda',
        'Rafael Uribe Uribe', 'San Cristóbal', 'Santa Fe',
        'Suba', 'Sumapaz', 'Teusaquillo', 'Tunjuelito',
        'Usaquén', 'Usme'
    ]
    return JsonResponse({'localidades': sorted(localidades)})


def api_barrios(request):
    localidad = request.GET.get('localidad', '')

    barrios = {
        'Usaquén': [
            'Barrancas','Bella Suiza','Bosque Medina','Brizos del Norte',
            'Cedritos','Cedritos Oriental','Cerros de Sotileza','Chicó Norte',
            'Chicó Norte II','Chicó Norte III','Chicó Reservado','Contador',
            'Country Club','El Cedro','El Refugio','El Verbenal','Horizontes',
            'La Cita','La Clarita','La Floresta Norte','La Frontera','La Punta',
            'La Uribe','Las Margaritas','Los Cedros','Los Cedros Oriental',
            'Los Cedros Occidental','Multicentro','Niza Norte','Orquídeas',
            'Paseo del Country','Pradera Norte','San Antonio Norte',
            'San Cristóbal Norte','Santa Ana','Santa Ana Occidental',
            'Santa Ana Oriental','Santa Bárbara','Santa Bárbara Central',
            'Santa Bárbara Occidental','Santa Bárbara Oriental','Soratama',
            'Torca','Toberín','Usaquén','Verbenal','Villa del Prado',
            'Villa Nidia','Vista Hermosa Norte',
        ],
        'Chapinero': [
            'Alcázares Norte','Bosque Calderón','Bosque Calderón Tejada',
            'Chapinero Alto','Chapinero Central','Chapinero Norte',
            'Chicó','Chicó Alto','Chicó Lago','Chicó Lago Occidental',
            'Chicó Reservado','El Castillo','El Paraíso','El Retiro',
            'Granada','Iberia','Juan XXIII','La Cabrera','La Salle',
            'Las Acacias','Las Nieves','Laureles','Los Olivos','Lourdes',
            'Marly','Meissen','Modelo Norte','Muequetá','Navarra',
            'Nueva Autopista','Palacio','Paraíso','Pardo Rubio',
            'Quinta Camacho','Refugio','Rosales','San Isidro Patios',
            'San Luis','San Martín','Sucre','Tejada','Tercer Milenio',
            'Unir I','Unir II',
        ],
        'Santa Fe': [
            'Atanasio Girardot','Bello Horizonte','Belén','Bosque Izquierdo',
            'Egipto','El Guavio','Germania','La Alameda','La Candelaria',
            'La Concordia','La Macarena','Las Aguas','Las Cruces',
            'Las Nieves','Lourdes','Miradores del Muelle','Paloquemao',
            'Paseo Los Libertadores','Ramírez','San Bernardo','San Diego',
            'San Victorino','Santa Bárbara','Santa Inés','Sevilla',
            'Veraguas','Veraguas Central',
        ],
        'San Cristóbal': [
            'Altamira','Arrayanes','Bello Horizonte','Buenos Aires',
            'Córdoba','El Quindío','El Rodeo','El Sosiego','El Triángulo',
            'Esfuerzos Unidos','Guacamayas','La Belleza','La Gloria',
            'La Hortúa','La Victoria','Las Malvinas','Los Alpes',
            'Los Pinos','Luna Park','Manantial','Montebello','Nariño Sur',
            'Nuevo Horizonte','Primero de Mayo','Rafael Uribe',
            'Ramajal','San Blas','San Cristóbal','San Isidro',
            'San Pedro','Santa Inés','Sosiego','Treinta y Un Sur',
            'Veinte de Julio','Villa del Cerro',
        ],
        'Usme': [
            'Alfonso López','Almendra','Armenia','Arrayanes',
            'Betania','Bosques de Bolonia','Brasil','Brazuelos',
            'Chuniza','Ciudad Usme','Comuneros','Compostela',
            'Danubio','El Bosque','El Danubio','El Destino',
            'El Portal','El Tuno','Fiscala','Gran Yomasa',
            'Granjas de San Pablo','Isabel Lleras','La Fiscala',
            'La Flora','La Reforma','Las Violetas','Los Comuneros',
            'Marichuela','Monteblanco','Nuevo Progreso',
            'Parque Entrenubes','Parques de Usme','Pedregal',
            'Porvenir','Puerta al Llano','San Andres',
            'San Isidro','San Libardo','Santa Librada',
            'Tocaimita','Usminia','Villa Alemania','Villa Betania',
            'Villa Consta','Villa Diana','Villa Emilia',
            'Villa Gloria','Yomasa',
        ],
        'Tunjuelito': [
            'Abraham Lincoln','Almendros','Babel','Bello Horizonte',
            'El Campin Sur','El Espino','El Meissen','El Mochuelo',
            'El Perdomo','El Tunal','Fátima','La Fiscala',
            'La Picota','La Quisqueya','Las Colinas','Laguneta',
            'Meissen','Muzú','Nuevo Muzú','Parque Entrenubes',
            'Petroleum','Quiroga','Rafael Uribe','San Benito',
            'San Carlos','Samore','Tunjuelito','Venecia',
            'Villa Gladys','Villa Italia',
        ],
        'Bosa': [
            'Abraham Lincoln','Alameda','Alfonso López','Apogeo',
            'Bella Flor','Bosa','Bosa Central','Bosa Occidental',
            'Brasil','Canoas','Carlos Albán','Casa Grande',
            'Ciudadela El Recreo','Clarelandia','El Eucaliptal',
            'El Porvenir','El Tintal','El Triunfo','El Vapor',
            'Escocia','Estación','Germania','Gran Britalia',
            'Gustavo Restrepo','Honduras','Independencia',
            'Israel','José Antonio Galán','José María Carbonell',
            'La Amistad','La Estancia','La Libertad','La Paz',
            'Las Margaritas','Laureles','Libertad','Llano Oriental',
            'Los Almendros','Metrovivienda','Naranjos','Niza Sur',
            'Nueva Colombia','Nuevo México','Olarte','Orlando Lara',
            'Pablo Neruda','Palermo Sur','Perdomo','Piamonte',
            'Portal de Bosa','Primavera','Recreo','San Bernardino',
            'San Diego','San Eugenio','San Jorge','San José',
            'San Pablo','Santiago de las Atalayas','Santo Domingo',
            'Tintal Norte','Tintal Sur','Villa Claudia','Villa del Río',
            'Villa Emma','Villa Nelly','Vista Hermosa',
        ],
        'Kennedy': [
            'Alquería','Américas','Andes Sur','Antonia Santos',
            'Banderas','Bavaria','Britalia','Calarcá',
            'Carlos Albán','Castilla','Catamarca','Ciudad Kennedy',
            'Ciudad Kennedy Central','Ciudad Kennedy Norte',
            'Ciudad Kennedy Occidental','Ciudad Kennedy Oriental',
            'Ciudad Kennedy Sur','Carvajal','El Amparo',
            'El Arriero','El Claret','El Jazmín','El Listón',
            'El Tingua','El Vergel','Galán','Granjas de Techo',
            'Hipódromo','Independencia','Jazmín','Kennedy',
            'La Alameda','La Alquería','La América','La Igualdad',
            'Las Margaritas','Las Palmeras','Llano Grande',
            'Los Álamos','Los Andes','Lusitania','María Paz',
            'Marsella','Milán','Nuevo Kennedy','Patio Bonito',
            'Perpetuo Socorro','Primavera','Quirigua Sur',
            'Restrepo','San Rafael','Santa Rosita','Santander',
            'Serranías','Timiza','Tintal','Tinto','Tunjuelito',
            'Valladolid','Villa Alsacia','Villa de la Torre',
            'Villa El Dorado','Villa Nelly','Villamar',
        ],
        'Fontibón': [
            'Aeropuerto','Bahía','Bavaria','Bogotá Cundinamarca',
            'Capellanía','Ciudad Salitre Occidental','Ciudad Salitre Oriental',
            'El Muelle','El Recuerdo','El Recuerdo Sur','Emilio Cifuentes',
            'Fontibón','Fontibón Centro','Fontibón San Pablo',
            'Granjas de Techo','Guadual','Kasandra','La Aldea',
            'La Cabaña','La Estancia','La Loma','Las Viñas',
            'Llanitos','Lusitania','Maravillas','Modelia',
            'Modelia Occidental','Modelia Oriental','Montana',
            'Murillo','Prados de Santa Bárbara','Prosperidad',
            'Puerto Horizonte','Rectángulo','Rinconada',
            'San Pablo','Tintal','Tintal Central','Tintal Norte',
            'Tintal Sur','Versalles','Villemar',
        ],
        'Engativá': [
            'Álamos','Álamos Norte','Bachué','Bolivia',
            'Bonanza','Boyacá Real','Brasil','Ciudadela Cafam',
            'Cortijo','Delia','El Muelle','El Palmar',
            'El Rubí','Engativá','Florencia','Francia',
            'Garcés Navas','Garces Navas','Guadalupe','Jardín Botánico',
            'La Cabañita','La Palma','La Pastora','Las Ferias',
            'Las Mercedes','Las Orquídeas','Lombardía','Los Ángeles',
            'Los Cerezos','Luis Carlos Galán','Manuela Beltrán',
            'Minuto de Dios','Morato','Morisco','Navarra',
            'Nueva Esperanza','Palermo','Primavera','Quirigua',
            'Quirigua Norte','Santa Cecilia','Santa Helenita',
            'Santa María del Lago','Sevilla','Tabora',
            'Villa Amalia','Villa Claver','Villa del Mar',
            'Villa Gladys','Villa Luz','Villamaría',
        ],
        'Suba': [
            'Alhambra','Aures','Aures I','Aures II',
            'Bilbao','Britalia','Buenavista','Casa Blanca Suba',
            'Casablanca','Cataluña','Ciudadela Colsubsidio',
            'Compartir','El Prado','El Rincón','Fontanar del Río',
            'Granada Norte','Guicán','Guiparma','Iberia',
            'Jardín de La Esperanza','La Academia','La Alborada',
            'La Campiña','La Conejera','La Esperanza Norte',
            'La Gaitana','La Isabela','La Palma','La Toscana',
            'Las Mercedes','Laureles','Lisboa','Lombardía',
            'Los Arrayanes','Los Naranjos','Mazurén','Niza',
            'Niza Norte','Niza Suba','Nueva Zelandia','Palonegro',
            'Pasadena','Pinar del Río','Potosí','Prado Pinzón',
            'Prado Veraniego','Prado Veraniego Norte',
            'Prado Veraniego Sur','Reserva','Rincón',
            'San Cayetano','San José del Prado','San Pedro de Tibabuyes',
            'Santa Bárbara Occidental','Santa Isabel','Suba',
            'Tibabita','Tibabuyes','Toscana','Villa Cindy',
            'Villa del Prado','Villa Elisa','Villa Hermosa',
        ],
        'Barrios Unidos': [
            '12 de Octubre','Alcázares','Andes Norte','Argentina',
            'Barrios Unidos','Benjamín Herrera','Boyacá','Campincito',
            'Colombia','Comuneros','Coveñas','Doce de Octubre',
            'El Campín','El Rosario','El Triunfo','Gaitán',
            'Jorge Eliécer Gaitán','José Martí','La Castellana',
            'La Esmeralda','La Florida','La Paz','La Patria',
            'Los Andes','Metrópolis','Minuto de Dios','Modelo',
            'Muequetá','Nicolás de Federman','Once de Noviembre',
            'Pablo VI','Palermo','Polo Club','Popular','Pradera',
            'Rionegro','Rioseco','San Fernando','Siete de Agosto',
            'Triangulo','Trinidad Galán','Unión','Veraguas',
        ],
        'Teusaquillo': [
            'Acevedo Tejada','Alameda','Américas','Armenia',
            'Campín','Capri','El Recuerdo','Espartillal',
            'Galerías','Gorgonzola','Hipódromo','La Esmeralda',
            'La Magdalena','La Soledad','Los Alcázares',
            'Maravillas','Marsella','Normandía','Palermo',
            'Parque Nacional','Quinta Paredes','Recuerdo',
            'Sagrado Corazón','San Luis','Santa Fe de Bogotá',
            'Soledad','Teusaquillo','Tibaitata',
        ],
        'Los Mártires': [
            'Colseguros','Eduardo Santos','El Listón','El Vergel',
            'Estación Central','La Favorita','La Pepita','La Sabana',
            'Las Cruces','Laches','Los Mártires','Paloquemao',
            'Ricaurte','San Victorino','Santa Isabel','Santafé',
            'Voto Nacional',
        ],
        'Antonio Nariño': [
            'Antonio Nariño','Ciudad Jardín Sur','Ciudad Berna',
            'Colombia','Cinco Huecos','El Claret','El Eden',
            'El Inglés','El Porvenir','Fatima','La Fragua',
            'Libertador','Muzú','Policarpa','Restrepo',
            'San Antonio','Santa Lucía','Santander',
        ],
        'Puente Aranda': [
            'Alcázares Sur','Alquería','Américas Occidental',
            'Autopista Sur','Batallon Caldas','Camelia',
            'Ciudad Montes','Ciudad Nariño','Colombia',
            'El Rosario','El Salitre','Galán','Jazmín',
            'La Asunción','La Camelia','La Favorita',
            'La Trinidad','Las Américas','Las Delicias',
            'Madelena','Ortezal','Pensilvania','Puente Aranda',
            'Quirigua','Salazar Gómez','Salitre','San Rafael',
            'Santa Matilde','Siete de Agosto','Trinidad',
            'Zona Industrial',
        ],
        'La Candelaria': [
            'Belén','Catedral','Egipto','La Candelaria',
            'Las Aguas','Las Nieves','Santa Bárbara',
        ],
        'Rafael Uribe Uribe': [
            'Bravo Páez','Chircales','Ciudad Jardín Sur',
            'Claret','Consuelo','Corinto','Diana Turbay',
            'El Mochuelo','El Triunfo','Granjas de San Pablo',
            'La Chucua','La Colina','La Flora','La Picota',
            'Laguneta','Llano Oriental','Los Alpes','Marruecos',
            'Marco Fidel Suárez','Molinos','Molinos Norte',
            'Molinos Sur','Olaya','Palermo Sur','Parque Entrenubes',
            'Quiroga','Rafael Uribe','San Agustín','San José',
            'Santa Lucía','Tesoro','Tibanica','Valles de Cafam',
            'Villa del Cerro',
        ],
        'Ciudad Bolívar': [
            'Arborizadora','Arborizadora Alta','Arborizadora Baja',
            'Bella Flor','Canoas','Casa de Teja','Compartir',
            'Danubio Azul','El Espino','El Mochuelo','El Paraíso',
            'El Tesoro','El Triunfo','Escocia','Estrella del Sur',
            'Jerusalén','Juan José Rondón','La Estancia',
            'La Esmeralda','La Gloria','La Isla','La Libertad',
            'La Paz','Las Brisas','Las Manitas','Laguneta',
            'Limonal','Lucero','Lucero Alto','Lucero Bajo',
            'Meissen','México','Minas','Monte Blanco',
            'Monteblanco','Perdomo','Perdomo Alto','Perdomo Bajo',
            'Potosí','Quiba','Quiba Alta','Quiba Baja',
            'Quebradaseca','San Francisco','San Isidro',
            'San Joaquín','San Pedro','Santa Rosita','Sierra Morena',
            'Tesoro','Tres Reyes','Tres Reyes Sur','Verona',
            'Villa Gloria','Vista Hermosa',
        ],
        'Sumapaz': [
            'Betania','Capitolio','Concepción','Erasmo',
            'La Unión','Las Vegas','Los Ríos','Nazareth',
            'Nueva Granada','Raizal','San Antonio','San Juan',
            'Santa Rosa',
        ],
    }

    if not localidad:
        return JsonResponse({'barrios': []})

    lista = barrios.get(localidad, [])
    return JsonResponse({'barrios': sorted(lista)})

@login_required
def eliminar_equipo(request, equipo_id):
    equipo   = get_object_or_404(Equipo, id=equipo_id)
    es_admin = request.user.rol == 'ADMIN'
    es_dueno = (
        request.user.rol == 'ENTRENADOR' and
        hasattr(request.user, 'entrenador') and
        equipo.entrenador == request.user.entrenador
    )

    if not es_admin and not es_dueno:
        messages.error(request, 'No tienes permiso para eliminar este equipo.')
        return redirect('inscripciones:mi_equipo')

    if request.method == 'POST':
        if es_admin:
            if equipo.eliminar_programada_para:
                messages.error(request, MENSAJE_ELIMINACION_PROGRAMADA)
                return redirect('inscripciones:lista_equipos')

            motivo = request.POST.get('motivo_eliminacion', '')
            if not motivo:
                motivo = _construir_motivo(
                    request.POST.get('motivo_tipo_eliminacion'),
                    request.POST.get('motivo_eliminacion_manual'),
                    MOTIVOS_ELIMINACION_EQUIPO,
                )
            _agendar_eliminacion_equipo(request, equipo, motivo)
            return redirect('inscripciones:lista_equipos')

        if es_dueno:
            if equipo.eliminar_programada_para:
                messages.error(request, MENSAJE_ELIMINACION_PROGRAMADA)
                return redirect('inscripciones:mi_equipo')
            nombre = equipo.nombre
            equipo.delete()
            messages.success(request, f'Equipo "{nombre}" eliminado correctamente.')
            return redirect('dashboard_entrenador')

    return render(request, 'inscripciones/confirmar_eliminar_equipo.html', {
        'equipo': equipo
    })


@login_required
def lista_equipos(request):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    nombre      = request.GET.get('nombre', '').strip()
    localidad   = request.GET.get('localidad', '').strip()
    estado      = request.GET.get('estado', '').strip()
    eliminacion = request.GET.get('eliminacion', '').strip()

    equipos = Equipo.objects.all().order_by('-fecha_registro')
    if nombre:
        equipos = equipos.filter(_nombre__icontains=nombre)
    if localidad:
        equipos = equipos.filter(_localidad__icontains=localidad)
    if estado:
        equipos = equipos.filter(_estado=estado)
    if eliminacion == 'si':
        equipos = equipos.filter(_eliminar_programada_para__isnull=False)

    export = request.GET.get('export', '').strip().lower()
    equipos_export = list(equipos.select_related('entrenador'))
    if export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Equipos'
        ws.append(['Reporte de equipos y entrenadores F.A.S'])
        ws.append([f'Fecha: {date.today().strftime("%d/%m/%Y")}'])
        ws.append([f'Filtros: nombre={nombre or "todos"}; localidad={localidad or "todas"}; estado={estado or "todos"}; eliminacion={eliminacion or "todos"}'])
        ws.append([])
        ws.append([
            'Equipo', 'Estado', 'Categoria', 'Localidad', 'Barrio', 'Fundacion',
            'Entrenador', 'Documento entrenador', 'Correo entrenador', 'Telefono entrenador',
            'Experiencia', 'Eliminacion programada', 'Motivo eliminacion', 'Motivo rechazo'
        ])
        header_row = ws.max_row
        for equipo in equipos_export:
            ws.append([
                equipo.nombre,
                equipo.estado_display,
                equipo.categoria_display,
                equipo.localidad,
                equipo.barrio,
                equipo.anio_fundacion,
                f'{equipo.entrenador.nombres} {equipo.entrenador.apellidos}',
                equipo.entrenador.num_documento,
                equipo.entrenador.email,
                equipo.entrenador.telefono,
                equipo.entrenador.experiencia,
                equipo.eliminar_programada_para.strftime('%d/%m/%Y %H:%M') if equipo.eliminar_programada_para else '',
                equipo.motivo_eliminacion or '',
                equipo.motivo_rechazo or '',
            ])
        ws.freeze_panes = f'A{header_row + 1}'
        for cell in ws[header_row]:
            font = copy(cell.font)
            fill = copy(cell.fill)
            font.bold = True
            font.color = 'FFFFFF'
            fill.fill_type = 'solid'
            fill.fgColor = '0D47A1'
            cell.font = font
            cell.fill = fill
        for column in ws.columns:
            letter = column[0].column_letter
            ws.column_dimensions[letter].width = min(34, max(12, max(len(str(cell.value or '')) for cell in column) + 2))
        output = io.BytesIO()
        wb.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_equipos_entrenadores.xlsx"'
        return response

    if export == 'pdf':
        response = HttpResponse(_generar_pdf_equipos(equipos_export, {
            'nombre': nombre,
            'localidad': localidad,
            'estado': estado,
            'eliminacion': eliminacion,
        }), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_equipos_entrenadores.pdf"'
        return response
    
    paginator = Paginator(equipos, 12)
    page      = request.GET.get('page')
    equipos   = paginator.get_page(page)

    return render(request, 'inscripciones/lista_equipos.html', {
        'equipos':   equipos,
        'nombre':      nombre,
        'localidad':   localidad,
        'estado':      estado,
        'eliminacion': eliminacion,
        'estados':   Equipo.Estado.choices,
        'motivos_rechazo': MOTIVOS_RECHAZO_EQUIPO,
        'motivos_eliminacion': MOTIVOS_ELIMINACION_EQUIPO,
        'total': paginator.count,
    })

@login_required
def aprobar_equipo(request, equipo_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    equipo = get_object_or_404(Equipo, id=equipo_id)
    accion = request.POST.get('accion')

    if accion == 'aprobar':
        equipo.estado         = Equipo.Estado.APROBADO
        equipo.motivo_rechazo = None  # ← limpiar si había uno previo
        equipo.fecha_rechazo = None
        equipo.bloqueado_hasta = None
        equipo.eliminar_programada_para = None
        messages.success(request, f'Equipo "{equipo.nombre}" aprobado.')

    elif accion == 'rechazar':
        motivo = request.POST.get('motivo', '').strip()
        if not motivo:
            motivo = _construir_motivo(
                request.POST.get('motivo_tipo'),
                request.POST.get('motivo_manual'),
                MOTIVOS_RECHAZO_EQUIPO,
            )
        if not motivo:
            messages.error(request, 'Debes indicar el motivo del rechazo.')
            return redirect('inscripciones:lista_equipos')
        equipo.estado         = Equipo.Estado.RECHAZADO
        equipo.motivo_rechazo = motivo
        equipo.fecha_rechazo = timezone.now()
        equipo.bloqueado_hasta = timezone.now() + timedelta(days=15)
        messages.error(request, f'Equipo "{equipo.nombre}" rechazado.')

    equipo.save()
    return redirect('inscripciones:lista_equipos')


@login_required
def programar_eliminacion_equipo(request, equipo_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    equipo = get_object_or_404(Equipo, id=equipo_id)

    if request.method != 'POST':
        return redirect('inscripciones:lista_equipos')

    motivo = request.POST.get('motivo_eliminacion', '')
    if not motivo:
        motivo = _construir_motivo(
            request.POST.get('motivo_tipo_eliminacion'),
            request.POST.get('motivo_eliminacion_manual'),
            MOTIVOS_ELIMINACION_EQUIPO,
        )
    _agendar_eliminacion_equipo(request, equipo, motivo)
    return redirect('inscripciones:lista_equipos')

@login_required
def lista_jugadores(request):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')
    
    equipo = _get_equipo_entrenador(request)
    if not equipo:
        messages.error(request, 'No tienes un equipo registrado.')
        return redirect ('inscripciones:mi_equipo')
    
    jugadores = Jugador.objects.filter(equipo = equipo).order_by('_dorsal')
    total_jugadores = jugadores.count()
    cupo_maximo_jugadores = Jugador.MAX_JUGADORES_POR_EQUIPO
    cupos_disponibles = max(cupo_maximo_jugadores - total_jugadores, 0)

    return render(request, 'inscripciones/lista_jugadores.html', {
        'equipo': equipo,
        'equipos': equipos_del_entrenador(request.user),
        'jugadores': jugadores,
        'total_jugadores': total_jugadores,
        'cupo_maximo_jugadores': cupo_maximo_jugadores,
        'cupos_disponibles': cupos_disponibles,
        'equipo_lleno': cupos_disponibles == 0,
    })

@login_required
def registrar_jugador(request):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    equipo = _get_equipo_entrenador(request)
    if not equipo:
        messages.error(request, 'Debes registrar un equipo primero.')
        return redirect('inscripciones:mi_equipo')

    if equipo.estado != 'APROBADO':
        messages.error(request, 'Tu equipo debe estar aprobado para registrar jugadores.')
        return redirect('inscripciones:mi_equipo')

    ahora = timezone.now()
    if equipo.eliminar_programada_para and equipo.eliminar_programada_para > ahora:
        messages.error(request, 'Este equipo tiene una eliminación programada y no puede registrar jugadores.')
        return redirect('inscripciones:mi_equipo')

    total_jugadores = Jugador.objects.filter(equipo=equipo).count()
    if total_jugadores >= Jugador.MAX_JUGADORES_POR_EQUIPO:
        messages.error(
            request,
            f'Tu equipo ya alcanzó el máximo de {Jugador.MAX_JUGADORES_POR_EQUIPO} jugadores.'
        )
        return redirect('inscripciones:lista_jugadores')

    form = RegistroJugadorForm(request.POST or None, equipo=equipo)

    if request.method == 'POST' and form.is_valid():
        data = form.cleaned_data
        try:
            jugador = Jugador()
            jugador.nombres          = data['nombres']
            jugador.apellidos        = data['apellidos']
            jugador.num_documento    = data['num_documento']
            jugador.fecha_nacimiento = data['fecha_nacimiento']
            jugador.email            = data['email']
            jugador.telefono         = data['telefono']
            jugador._rol             = Usuario.Roles.JUGADOR
            jugador.dorsal           = data['dorsal']
            jugador.pie_dominante    = data['pie_dominante']
            jugador.posicion         = data['posicion']
            jugador.equipo           = equipo
            jugador.set_password(data['password'])
            jugador.save()

            _enviar_credenciales_jugador(jugador, data['password'], request)

            messages.success(request, f'Jugador {data["nombres"]} registrado correctamente.')
            return redirect('inscripciones:lista_jugadores')

        except IntegrityError as e:
            # Mapear la constraint violada al campo correspondiente
            error = str(e).lower()
            if 'email' in error or '_email' in error:
                form.add_error('email', 'Este correo ya está registrado.')
            elif 'num_documento' in error or '_num_documento' in error:
                form.add_error('num_documento', 'Este documento ya está registrado.')
            elif 'telefono' in error or '_telefono' in error:
                form.add_error('telefono', 'Este teléfono ya está registrado.')
            elif 'dorsal' in error or '_dorsal' in error:
                form.add_error('dorsal', f'El dorsal ya está en uso en este equipo.')
            else:
                form.add_error(None, f'Error de integridad en base de datos: {e}')

        except ValueError as e:
            form.add_error(None, str(e))

    return render(request, 'inscripciones/registrar_jugador.html', {
        'form': form,
        'equipo': equipo,
    })

@login_required
def eliminar_jugador(request, jugador_id):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    jugador = get_object_or_404(Jugador, id=jugador_id)
    equipo  = _get_equipo_entrenador(request)

    if jugador.equipo != equipo:
        messages.error(request, 'No tienes permiso para eliminar este jugador.')
        return redirect('inscripciones:lista_jugadores')

    if request.method == 'POST':
        nombre = f'{jugador.nombres} {jugador.apellidos}'
        jugador.delete()
        messages.success(request, f'Jugador "{nombre}" eliminado.')
        return redirect('inscripciones:lista_jugadores')

    return render(request, 'inscripciones/confirmar_eliminar_jugador.html', {
        'jugador': jugador
    })

@login_required
def carga_masiva_jugadores(request):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    equipo = _get_equipo_entrenador(request)
    if not equipo:
        messages.error(request, 'Debes registrar un equipo primero.')
        return redirect('inscripciones:mi_equipo')

    if equipo.estado != 'APROBADO':
        messages.error(request, 'Tu equipo debe estar aprobado para registrar jugadores.')
        return redirect('inscripciones:mi_equipo')

    ahora = timezone.now()
    if equipo.eliminar_programada_para and equipo.eliminar_programada_para > ahora:
        messages.error(request, 'Este equipo tiene una eliminación programada y no puede registrar jugadores.')
        return redirect('inscripciones:mi_equipo')

    total_jugadores = Jugador.objects.filter(equipo=equipo).count()
    if total_jugadores >= Jugador.MAX_JUGADORES_POR_EQUIPO:
        messages.error(
            request,
            f'Tu equipo ya alcanzó el máximo de {Jugador.MAX_JUGADORES_POR_EQUIPO} jugadores.'
        )
        return redirect('inscripciones:lista_jugadores')

    form = CargaMasivaJugadoresForm(request.POST or None, request.FILES or None)
    errores  = []
    exitosos = 0

    if request.method == 'POST' and form.is_valid():
        archivo = form.cleaned_data['archivo']
        nombre  = archivo.name.lower()

        try:
            if nombre.endswith('.xlsx'):
                filas = _leer_excel(archivo)
            else:
                filas = _leer_csv(archivo)

            exitosos, errores = _procesar_carga_masiva_jugadores(filas, equipo, request)

            if exitosos:
                messages.success(request, f'{exitosos} jugador(es) registrado(s) correctamente.')
            if errores:
                messages.error(request, f'{len(errores)} fila(s) con errores.')

        except Exception as e:
            messages.error(request, f'Error procesando el archivo: {str(e)}')

    return render(request, 'inscripciones/carga_masiva_jugadores.html', {
        'form':    form,
        'errores': errores,
        'equipo':  equipo,
    })

def _leer_excel(archivo):
    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, None)
        if not headers_row:
            return []

        headers = [str(h).strip().lower() if h else '' for h in headers_row]
        filas = []
        for row in rows_iter:
            if any(cell not in (None, '') for cell in row):
                filas.append(dict(zip(headers, row)))
        return filas
    finally:
        wb.close()


def _leer_csv(archivo):
    content = archivo.read().decode('utf-8-sig', errors='replace')
    reader  = csv.DictReader(io.StringIO(content))
    return [
        {(k or '').strip().lower(): v for k, v in row.items()}
        for row in reader
    ]


def _obtener_valor_fila(fila, keys):
    for key in keys:
        value = fila.get(key)
        if value not in (None, ''):
            return str(value).strip()
    return ''


def _parsear_fecha_nacimiento(fecha_val):
    if isinstance(fecha_val, datetime):
        return fecha_val.date(), None
    if isinstance(fecha_val, date):
        return fecha_val, None

    if isinstance(fecha_val, (int, float)):
        try:
            fecha_excel = from_excel(fecha_val)
            if isinstance(fecha_excel, datetime):
                return fecha_excel.date(), None
            if isinstance(fecha_excel, date):
                return fecha_excel, None
        except Exception:
            pass

    if isinstance(fecha_val, str):
        fecha_texto = fecha_val.strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(fecha_texto, fmt).date(), None
            except ValueError:
                continue
        return None, f'Fecha inválida "{fecha_texto}"'

    return None, 'Fecha inválida o vacía'


def _normalizar_fila_jugador(fila, equipo, num_fila):
    nombres = _obtener_valor_fila(fila, ['nombres', 'nombre'])
    apellidos = _obtener_valor_fila(fila, ['apellidos', 'apellido'])
    num_documento = _obtener_valor_fila(fila, ['num_documento', 'documento', 'cedula'])
    fecha_raw = fila.get('fecha_nacimiento') or fila.get('fecha nacimiento') or fila.get('nacimiento')
    email = _obtener_valor_fila(fila, ['email', 'correo']).lower()
    telefono = _obtener_valor_fila(fila, ['telefono', 'teléfono', 'celular'])
    dorsal_str = _obtener_valor_fila(fila, ['dorsal', 'numero', 'número'])
    pie_dominante = _obtener_valor_fila(fila, ['pie_dominante', 'pie dominante', 'pie']).lower()
    posicion = _obtener_valor_fila(fila, ['posicion', 'posición']).lower()
    password = _obtener_valor_fila(fila, ['password', 'contraseña', 'clave']) or 'Fas2024*'

    def _first_error(validators):
        for validator in validators:
            msg = validator()
            if msg:
                return msg
        return None

    def _validar_texto(valor, msg_vacio):
        texto = valor.strip()
        if not texto:
            return msg_vacio
        if len(texto) < 2:
            return 'Mínimo 2 caracteres.'
        if not re.match(r'^[a-záéíóúüñA-ZÁÉÍÓÚÜÑ\s]+$', texto):
            return 'Solo letras y espacios.'
        return None

    def _validar_regex(valor, patron, msg):
        return None if re.match(patron, valor) else msg

    def _validar_password(valor):
        reglas = (
            (len(valor) < 8, 'Mínimo 8 caracteres.'),
            (not re.search(r'[A-Z]', valor), 'Debe incluir una mayúscula.'),
            (not re.search(r'[0-9]', valor), 'Debe incluir un número.'),
            (not re.search(r'[^a-zA-Z0-9]', valor), 'Debe incluir un carácter especial.'),
        )
        for condicion, mensaje in reglas:
            if condicion:
                return mensaje
        return None

    def _parsear_dorsal(valor):
        try:
            dorsal_val = int(float(valor))
        except (TypeError, ValueError):
            return None, f'Dorsal inválido "{valor}"'
        if not 1 <= dorsal_val <= 99:
            return None, 'El dorsal debe estar entre 1 y 99'
        return dorsal_val, None

    def _validar_fecha(fecha):
        hoy = date.today()
        if fecha > hoy:
            return 'La fecha no puede ser futura.'
        edad = hoy.year - fecha.year - ((hoy.month, hoy.day) < (fecha.month, fecha.day))
        if edad > 100:
            return 'Fecha inválida.'
        return None

    campos_requeridos = {
        'nombres': nombres,
        'apellidos': apellidos,
        'num_documento': num_documento,
        'email': email,
        'telefono': telefono,
        'dorsal': dorsal_str,
        'pie_dominante': pie_dominante,
        'posicion': posicion,
    }
    faltantes = [k for k, v in campos_requeridos.items() if not v]
    if faltantes or fecha_raw in (None, ''):
        if fecha_raw in (None, '') and 'fecha_nacimiento' not in faltantes:
            faltantes.append('fecha_nacimiento')
        return None, f'Fila {num_fila}: Campos vacíos: {", ".join(faltantes)}'

    pies_validos = {c[0] for c in PIES if c[0]}
    posiciones_validas = {c[0] for c in POSICIONES if c[0]}
    error = _first_error([
        lambda: _validar_texto(nombres, 'El nombre es obligatorio.'),
        lambda: _validar_texto(apellidos, 'Los apellidos son obligatorios.'),
        lambda: _validar_regex(num_documento, r'^\d{6,12}$', 'Documento inválido. Entre 6 y 12 dígitos numéricos.'),
        lambda: _validar_regex(email, r'^[^\s@]+@[^\s@]+\.[^\s@]{2,}$', 'Ingresa un correo válido.'),
        lambda: _validar_regex(telefono, r'^3[0-9]{9}$', 'Número colombiano válido (ej: 3001234567).'),
        lambda: None if pie_dominante in pies_validos else 'Selecciona una opción.',
        lambda: None if posicion in posiciones_validas else 'Selecciona una posición.',
        lambda: _validar_password(password),
    ])
    if error:
        return None, f'Fila {num_fila}: {error}'

    dorsal, dorsal_error = _parsear_dorsal(dorsal_str)
    if dorsal_error:
        return None, f'Fila {num_fila}: {dorsal_error}'

    fecha_nacimiento, fecha_error = _parsear_fecha_nacimiento(fecha_raw)
    if fecha_error:
        return None, f'Fila {num_fila}: {fecha_error}'

    fecha_error = _validar_fecha(fecha_nacimiento)
    if fecha_error:
        return None, f'Fila {num_fila}: {fecha_error}'

    valido, msg = validar_edad_categoria(fecha_nacimiento, equipo.categoria)
    if not valido:
        return None, f'Fila {num_fila}: {msg}'

    return {
        'num_fila': num_fila,
        'nombres': nombres,
        'apellidos': apellidos,
        'num_documento': num_documento,
        'fecha_nacimiento': fecha_nacimiento,
        'email': email,
        'telefono': telefono,
        'dorsal': dorsal,
        'pie_dominante': pie_dominante,
        'posicion': posicion,
        'password': password,
    }, None


def _buscar_existentes_en_bloques(modelo, campo, valores, **filtros):
    if not valores:
        return set()

    existentes = set()
    valores_lista = list(valores)
    chunk_size = 800

    for i in range(0, len(valores_lista), chunk_size):
        bloque = valores_lista[i:i + chunk_size]
        qs = modelo.objects.filter(**filtros).filter(**{f'{campo}__in': bloque})
        existentes.update(qs.values_list(campo, flat=True))

    return existentes


def _procesar_carga_masiva_jugadores(filas, equipo, request):
    errores = []
    filas_limpias = []

    max_jugadores_equipo = Jugador.MAX_JUGADORES_POR_EQUIPO
    jugadores_actuales = Jugador.objects.filter(equipo=equipo).count()
    cupos_disponibles = max(max_jugadores_equipo - jugadores_actuales, 0)

    if cupos_disponibles <= 0:
        return 0, [
            f'El equipo ya alcanzó el máximo de {max_jugadores_equipo} jugadores.'
        ]

    emails_archivo = set()
    documentos_archivo = set()
    telefonos_archivo = set()
    dorsales_archivo = set()

    for num_fila, fila in enumerate(filas, start=2):
        data, error = _normalizar_fila_jugador(fila, equipo, num_fila)
        if error:
            errores.append(error)
            continue

        if data['email'] in emails_archivo:
            errores.append(f'Fila {num_fila}: Correo "{data["email"]}" duplicado en el archivo')
            continue
        if data['num_documento'] in documentos_archivo:
            errores.append(f'Fila {num_fila}: Documento "{data["num_documento"]}" duplicado en el archivo')
            continue
        if data['telefono'] in telefonos_archivo:
            errores.append(f'Fila {num_fila}: Teléfono "{data["telefono"]}" duplicado en el archivo')
            continue
        if data['dorsal'] in dorsales_archivo:
            errores.append(f'Fila {num_fila}: Dorsal {data["dorsal"]} duplicado en el archivo')
            continue

        emails_archivo.add(data['email'])
        documentos_archivo.add(data['num_documento'])
        telefonos_archivo.add(data['telefono'])
        dorsales_archivo.add(data['dorsal'])
        filas_limpias.append(data)

    if not filas_limpias:
        return 0, errores

    emails_existentes = _buscar_existentes_en_bloques(Usuario, '_email', emails_archivo)
    documentos_existentes = _buscar_existentes_en_bloques(Usuario, '_num_documento', documentos_archivo)
    telefonos_existentes = _buscar_existentes_en_bloques(Usuario, '_telefono', telefonos_archivo)
    dorsales_existentes = _buscar_existentes_en_bloques(Jugador, '_dorsal', dorsales_archivo, equipo=equipo)

    registros_guardados = []

    for data in filas_limpias:
        num_fila = data['num_fila']

        if len(registros_guardados) >= cupos_disponibles:
            errores.append(
                f'Fila {num_fila}: No hay cupos disponibles. Máximo {max_jugadores_equipo} jugadores por equipo'
            )
            continue

        if data['email'] in emails_existentes:
            errores.append(f'Fila {num_fila}: Correo "{data["email"]}" ya registrado')
            continue
        if data['num_documento'] in documentos_existentes:
            errores.append(f'Fila {num_fila}: Documento "{data["num_documento"]}" ya registrado')
            continue
        if data['telefono'] in telefonos_existentes:
            errores.append(f'Fila {num_fila}: Teléfono "{data["telefono"]}" ya registrado')
            continue
        if data['dorsal'] in dorsales_existentes:
            errores.append(f'Fila {num_fila}: Dorsal {data["dorsal"]} ya en uso')
            continue

        try:
            jugador = Jugador()
            jugador.nombres = data['nombres']
            jugador.apellidos = data['apellidos']
            jugador.num_documento = data['num_documento']
            jugador.fecha_nacimiento = data['fecha_nacimiento']
            jugador.email = data['email']
            jugador.telefono = data['telefono']
            jugador._rol = Usuario.Roles.JUGADOR
            jugador.dorsal = data['dorsal']
            jugador.pie_dominante = data['pie_dominante']
            jugador.posicion = data['posicion']
            jugador.equipo = equipo
            jugador.password = make_password(data['password'])
            jugador.save()

            registros_guardados.append((jugador, data['password']))

            emails_existentes.add(data['email'])
            documentos_existentes.add(data['num_documento'])
            telefonos_existentes.add(data['telefono'])
            dorsales_existentes.add(data['dorsal'])

        except IntegrityError:
            errores.append(f'Fila {num_fila}: Conflicto de datos duplicados al guardar')
        except ValueError as e:
            errores.append(f'Fila {num_fila}: {str(e)}')
        except Exception as e:
            errores.append(f'Fila {num_fila}: {str(e)}')

    if registros_guardados:
        enviar_credenciales_jugadores_lote(registros_guardados, request)

    return len(registros_guardados), errores

@login_required
def editar_jugador(request, jugador_id):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    jugador = get_object_or_404(Jugador, id=jugador_id)
    equipo  = _get_equipo_entrenador(request)

    if jugador.equipo != equipo:
        messages.error(request, 'No tienes permiso para editar este jugador.')
        return redirect('inscripciones:lista_jugadores')

    form = EditarJugadorEntrenadorForm(
        request.POST or None,
        jugador_pk=jugador.id,
        equipo=equipo,
        initial={
            'dorsal':        jugador.dorsal,
            'pie_dominante': jugador.pie_dominante,
            'posicion':      jugador.posicion,
        }
    )

    if request.method == 'POST':
        if form.is_valid():
            data = form.cleaned_data
            try:
                jugador.dorsal        = data['dorsal']
                jugador.pie_dominante = data['pie_dominante']
                jugador.posicion      = data['posicion']
                jugador.save()
                messages.success(request, 'Jugador actualizado correctamente.')
                return redirect('inscripciones:lista_jugadores')
            except ValueError as e:
                form.add_error(None, str(e))

    return render(request, 'inscripciones/editar_jugador.html', {
        'form':    form,
        'jugador': jugador,
        'equipo':  equipo,
    })

@login_required
def editar_perfil_jugador(request):
    if request.user.rol != 'JUGADOR':
        return redirect('dashboard_entrenador')

    user = request.user
    form = EditarPerfilJugadorForm(
        request.POST or None,
        jugador_pk=user.id,
        initial={
            'nombres':       user.nombres,
            'apellidos':     user.apellidos,
            'num_documento': user.num_documento,
            'email':         user.email,
            'telefono':      user.telefono,
        }
    )

    if request.method == 'POST':
        if form.is_valid():
            data = form.cleaned_data
            try:
                user.nombres       = data['nombres']
                user.apellidos     = data['apellidos']
                user.num_documento = data['num_documento']
                user.email         = data['email']
                user.telefono      = data['telefono']

                password_actual = data.get('password_actual')
                password_nueva  = data.get('password_nueva')
                if password_actual and password_nueva:
                    if user.check_password(password_actual):
                        user.set_password(password_nueva)
                        update_session_auth_hash(request, user)
                        messages.success(request, 'Contraseña actualizada.')
                    else:
                        form.add_error('password_actual', 'Contraseña actual incorrecta.')
                        return render(request, 'accounts/roles/editar_perfil_jugador.html', {'form': form})

                user.save()
                messages.success(request, 'Perfil actualizado correctamente.')
                return redirect('editar_jugador_perfil')

            except ValueError as e:
                form.add_error(None, str(e))
            except IntegrityError as e:
                err = str(e).lower()
                if 'email' in err:
                    form.add_error('email', 'Este correo ya está registrado.')
                elif 'documento' in err:
                    form.add_error('num_documento', 'Este documento ya está registrado.')
                elif 'telefono' in err:
                    form.add_error('telefono', 'Este teléfono ya está registrado.')

    return render(request, 'accounts/roles/editar_perfil_jugador.html', {'form': form})

@login_required
def lista_canchas(request):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    nombre       = request.GET.get('nombre', '').strip()
    localidad    = request.GET.get('localidad', '').strip()
    disciplina   = request.GET.get('disciplina', '').strip()
    disponibilidad = request.GET.get('disponibilidad', '').strip()

    canchas = Cancha.objects.all().order_by('_nombre_escenario')
    if nombre:
        canchas = canchas.filter(_nombre_escenario__icontains=nombre)
    if localidad:
        canchas = canchas.filter(_localidad__icontains=localidad)
    if disciplina:
        canchas = canchas.filter(_tipo_disciplina=disciplina)
    if disponibilidad:
        canchas = canchas.filter(_disponibilidad=disponibilidad)
    
    paginator = Paginator(canchas, 15)
    page      = request.GET.get('page')
    canchas   = paginator.get_page(page)

    canchas_mapa = Cancha.objects.filter(
        _latitud__isnull=False,
        _longitud__isnull=False
    ).values('_nombre_escenario', '_direccion_exacta', '_localidad',
             '_tipo_disciplina', '_disponibilidad', '_latitud', '_longitud', 'id')

    canchas_json = json.dumps([
        {
            'id':         c['id'],
            'nombre':     c['_nombre_escenario'],
            'direccion':  c['_direccion_exacta'],
            'localidad':  c['_localidad'],
            'disciplina': c['_tipo_disciplina'],
            'lat':        c['_latitud'],
            'lng':        c['_longitud'],
        }
        for c in canchas_mapa
    ])

    return render(request, 'inscripciones/canchas/lista_canchas.html', {
        'canchas':        canchas,
        'total': paginator.count,
        'nombre':         nombre,
        'localidad':      localidad,
        'disciplina':     disciplina,
        'disponibilidad': disponibilidad,
        'disciplinas':    Cancha.TipoDisciplina.choices,
        'disponibilidades': Cancha.Disponibilidad.choices,
        'canchas_json': canchas_json,
    })


# ── Crear cancha ──
@login_required
def crear_cancha(request):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    form = CanchaForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            data = form.cleaned_data
            try:
                cancha = Cancha()
                cancha.codigo_idrd            = data.get('codigo_idrd')
                cancha.nombre_escenario       = data['nombre_escenario']
                cancha.localidad              = data['localidad']
                cancha.barrio                 = data['barrio']
                cancha.direccion_exacta       = data['direccion_exacta']
                cancha.codigo_rupi            = data.get('codigo_rupi')
                cancha.tipo_disciplina        = data['tipo_disciplina']
                cancha.tipo_superficie        = data['tipo_superficie']
                cancha.medidas_area           = data['medidas_area']
                cancha.estado_conservacion    = data['estado_conservacion']
                cancha.tiene_iluminacion      = data.get('tiene_iluminacion', False)
                cancha.tiene_cerramiento      = data.get('tiene_cerramiento', False)
                cancha.capacidad_espectadores = data['capacidad_espectadores']
                cancha.observaciones_tecnicas = data.get('observaciones_tecnicas')
                lat, lng = geodificar_direccion(cancha.direccion_exacta)
                cancha.latitud = lat
                cancha.longitud = lng
                cancha.save()
                messages.success(request, f'Cancha "{cancha.nombre_escenario}" registrada correctamente.')
                return redirect('inscripciones:lista_canchas')
            except ValueError as e:
                form.add_error(None, str(e))

    return render(request, 'inscripciones/canchas/form_cancha.html', {
        'form':   form,
        'titulo': 'Registrar Cancha',
        'accion': 'Registrar',
    })


# ── Editar cancha ──
@login_required
def editar_cancha(request, cancha_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    cancha = get_object_or_404(Cancha, id=cancha_id)
    form   = CanchaForm(
        request.POST or None,
        cancha_pk=cancha_id,
        initial={
            'codigo_idrd':            cancha.codigo_idrd,
            'nombre_escenario':       cancha.nombre_escenario,
            'localidad':              cancha.localidad,
            'barrio':                 cancha.barrio,
            'direccion_exacta':       cancha.direccion_exacta,
            'codigo_rupi':            cancha.codigo_rupi,
            'tipo_disciplina':        cancha.tipo_disciplina,
            'tipo_superficie':        cancha.tipo_superficie,
            'medidas_area':           cancha.medidas_area,
            'estado_conservacion':    cancha.estado_conservacion,
            'tiene_iluminacion':      cancha.tiene_iluminacion,
            'tiene_cerramiento':      cancha.tiene_cerramiento,
            'capacidad_espectadores': cancha.capacidad_espectadores,
            'observaciones_tecnicas': cancha.observaciones_tecnicas,
        }
    )

    if request.method == 'POST':
        if form.is_valid():
            data = form.cleaned_data
            try:
                cancha.codigo_idrd            = data.get('codigo_idrd')
                cancha.nombre_escenario       = data['nombre_escenario']
                cancha.localidad              = data['localidad']
                cancha.barrio                 = data['barrio']
                cancha.direccion_exacta       = data['direccion_exacta']
                cancha.codigo_rupi            = data.get('codigo_rupi')
                cancha.tipo_disciplina        = data['tipo_disciplina']
                cancha.tipo_superficie        = data['tipo_superficie']
                cancha.medidas_area           = data['medidas_area']
                cancha.estado_conservacion    = data['estado_conservacion']
                cancha.tiene_iluminacion      = data.get('tiene_iluminacion', False)
                cancha.tiene_cerramiento      = data.get('tiene_cerramiento', False)
                cancha.capacidad_espectadores = data['capacidad_espectadores']
                cancha.observaciones_tecnicas = data.get('observaciones_tecnicas')
                lat, lng = geodificar_direccion(cancha.direccion_exacta)
                cancha.latitud = lat
                cancha.longitud = lng
                cancha.save()
                messages.success(request, 'Cancha actualizada correctamente.')
                return redirect('inscripciones:lista_canchas')
            except ValueError as e:
                form.add_error(None, str(e))

    return render(request, 'inscripciones/canchas/form_cancha.html', {
        'form':   form,
        'cancha': cancha,
        'titulo': 'Editar Cancha',
        'accion': 'Guardar cambios',
    })


# ── Eliminar cancha ──
@login_required
def eliminar_cancha(request, cancha_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    cancha = get_object_or_404(Cancha, id=cancha_id)

    if request.method == 'POST':
        nombre = cancha.nombre_escenario
        cancha.delete()
        messages.success(request, f'Cancha "{nombre}" eliminada.')
        return redirect('inscripciones:lista_canchas')

    return render(request, 'inscripciones/canchas/confirmar_eliminar_cancha.html', {
        'cancha': cancha
    })


# ── Cambiar disponibilidad ──
@login_required
def cambiar_disponibilidad(request, cancha_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    cancha = get_object_or_404(Cancha, id=cancha_id)

    if request.method == 'POST':
        nueva = request.POST.get('disponibilidad')
        try:
            cancha.disponibilidad = nueva
            cancha.save()
            messages.success(request, f'Disponibilidad actualizada a "{cancha.disponibilidad_display}".')
        except ValueError as e:
            messages.error(request, str(e))

    return redirect('inscripciones:lista_canchas')


# ── Carga masiva ──
@login_required
def carga_masiva_canchas(request):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    form     = CargaMasivaCanchasForm(request.POST or None, request.FILES or None)
    errores  = []
    exitosos = 0

    if request.method == 'POST' and form.is_valid():
        archivo = form.cleaned_data['archivo']
        nombre  = archivo.name.lower()
        try:
            if nombre.endswith('.xlsx'):
                filas = _leer_excel(archivo)
            else:
                filas = _leer_csv(archivo)

            for i, fila in enumerate(filas, start=2):
                resultado = _procesar_fila_cancha(fila, i)
                if resultado['ok']:
                    exitosos += 1
                else:
                    errores.append(resultado['error'])

            if exitosos:
                messages.success(request, f'{exitosos} cancha(s) registrada(s) correctamente.')
            if errores:
                messages.error(request, f'{len(errores)} fila(s) con errores.')

        except Exception as e:
            messages.error(request, f'Error procesando el archivo: {str(e)}')

    return render(request, 'inscripciones/canchas/carga_masiva_canchas.html', {
        'form':    form,
        'errores': errores,
    })


def _procesar_fila_cancha(fila, num_fila):
    try:
        def get(keys):
            for k in keys:
                if k in fila and fila[k] not in (None, ''):
                    return str(fila[k]).strip()
            return ''

        nombre_escenario      = get(['nombre_escenario', 'nombre'])
        localidad             = get(['localidad'])
        barrio                = get(['barrio'])
        direccion_exacta      = get(['direccion_exacta', 'direccion'])
        tipo_disciplina       = get(['tipo_disciplina', 'disciplina'])
        tipo_superficie       = get(['tipo_superficie', 'superficie'])
        medidas_area          = get(['medidas_area', 'medidas'])
        estado_conservacion   = get(['estado_conservacion', 'estado'])
        capacidad_str         = get(['capacidad_espectadores', 'capacidad']) or '0'
        codigo_idrd           = get(['codigo_idrd', 'codigo'])
        codigo_rupi           = get(['codigo_rupi', 'rupi'])
        tiene_iluminacion     = get(['tiene_iluminacion', 'iluminacion']).lower() in ('si', 'sí', 'true', '1', 'yes')
        tiene_cerramiento     = get(['tiene_cerramiento', 'cerramiento']).lower() in ('si', 'sí', 'true', '1', 'yes')
        observaciones         = get(['observaciones_tecnicas', 'observaciones'])

        campos_requeridos = {
            'nombre_escenario':    nombre_escenario,
            'localidad':           localidad,
            'barrio':              barrio,
            'direccion_exacta':    direccion_exacta,
            'tipo_disciplina':     tipo_disciplina,
            'tipo_superficie':     tipo_superficie,
            'medidas_area':        medidas_area,
            'estado_conservacion': estado_conservacion,
        }
        faltantes = [k for k, v in campos_requeridos.items() if not v]
        if faltantes:
            return {'ok': False, 'error': f'Fila {num_fila}: Campos vacíos: {", ".join(faltantes)}'}

        # Mapeo flexible de valores
        disciplina_map = {
            'futbol 11': 'FUTBOL_11', 'fútbol 11': 'FUTBOL_11',
            'futbol 8':  'FUTBOL_8',  'fútbol 8':  'FUTBOL_8',
            'futbol 5':  'FUTBOL_5',  'fútbol 5':  'FUTBOL_5',
            'microfutbol': 'MICROFUTBOL', 'microfútbol': 'MICROFUTBOL',
        }
        superficie_map = {
            'sintetica':      'SINTETICA',
            'sintética':      'SINTETICA',
            'natural':        'NATURAL',
            'arena':          'ARENA',
            'cemento':        'CEMENTO',
            'dura':           'CEMENTO',
            'cemento (dura)': 'CEMENTO',
            'cemento (duro)': 'CEMENTO',
        }
        conservacion_map = {
            'bueno': 'BUENO', 'regular': 'REGULAR',
            'malo':  'MALO',  'critico': 'CRITICO', 'crítico': 'CRITICO',
        }

        tipo_disciplina_val    = disciplina_map.get(tipo_disciplina.lower(), tipo_disciplina.upper())
        tipo_superficie_val    = superficie_map.get(tipo_superficie.lower(), tipo_superficie.upper())
        estado_conservacion_val = conservacion_map.get(estado_conservacion.lower(), estado_conservacion.upper())

        datos_cancha = {
            'codigo_idrd': codigo_idrd,
            'nombre_escenario': nombre_escenario,
            'localidad': localidad,
            'barrio': barrio,
            'direccion_exacta': direccion_exacta,
            'codigo_rupi': codigo_rupi,
            'tipo_disciplina': tipo_disciplina_val,
            'tipo_superficie': tipo_superficie_val,
            'medidas_area': medidas_area,
            'estado_conservacion': estado_conservacion_val,
            'tiene_iluminacion': tiene_iluminacion,
            'tiene_cerramiento': tiene_cerramiento,
            'capacidad_espectadores': int(float(capacidad_str)) if capacidad_str else 0,
            'observaciones_tecnicas': observaciones,
        }

        form = CanchaForm(datos_cancha)
        if not form.is_valid():
            errores = []
            for campo, mensajes in form.errors.items():
                for mensaje in mensajes:
                    if campo == '__all__':
                        errores.append(str(mensaje))
                    else:
                        errores.append(f'{campo}: {mensaje}')
            return {'ok': False, 'error': f'Fila {num_fila}: {"; ".join(errores)}'}

        cancha = Cancha()
        cancha.codigo_idrd            = form.cleaned_data.get('codigo_idrd')
        cancha.nombre_escenario       = form.cleaned_data['nombre_escenario']
        cancha.localidad              = form.cleaned_data['localidad']
        cancha.barrio                 = form.cleaned_data['barrio']
        cancha.direccion_exacta       = form.cleaned_data['direccion_exacta']
        cancha.codigo_rupi            = form.cleaned_data.get('codigo_rupi')
        cancha.tipo_disciplina        = form.cleaned_data['tipo_disciplina']
        cancha.tipo_superficie        = form.cleaned_data['tipo_superficie']
        cancha.medidas_area           = form.cleaned_data['medidas_area']
        cancha.estado_conservacion    = form.cleaned_data['estado_conservacion']
        cancha.tiene_iluminacion      = form.cleaned_data.get('tiene_iluminacion', False)
        cancha.tiene_cerramiento      = form.cleaned_data.get('tiene_cerramiento', False)
        cancha.capacidad_espectadores = form.cleaned_data['capacidad_espectadores']
        cancha.observaciones_tecnicas = form.cleaned_data.get('observaciones_tecnicas')
        lat, lng = geodificar_direccion(cancha.direccion_exacta)
        cancha.latitud = lat
        cancha.longitud = lng
        cancha.save()
        time.sleep(1)

        return {'ok': True}

    except Exception as e:
        return {'ok': False, 'error': f'Fila {num_fila}: {str(e)}'}


# ── Lista canchas entrenador — solo lectura ──
@login_required
def lista_canchas_entrenador(request):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    nombre     = request.GET.get('nombre', '').strip()
    localidad  = request.GET.get('localidad', '').strip()
    disciplina = request.GET.get('disciplina', '').strip()

    canchas = Cancha.objects.filter(
        _disponibilidad=Cancha.Disponibilidad.DISPONIBLE
    ).order_by('_nombre_escenario')

    if nombre:
        canchas = canchas.filter(_nombre_escenario__icontains=nombre)
    if localidad:
        canchas = canchas.filter(_localidad__icontains=localidad)
    if disciplina:
        canchas = canchas.filter(_tipo_disciplina=disciplina)
    
    todas_canchas = canchas

    paginator = Paginator(canchas, 12)
    page      = request.GET.get('page')
    canchas   = paginator.get_page(page)

    canchas_json = []
    for c in todas_canchas:  
        if c._latitud and c._longitud:
            canchas_json.append({
                "nombre": c._nombre_escenario,
                "direccion": c._direccion_exacta,
                "lat": c._latitud,
                "lng": c._longitud,
            })

    return render(request, 'inscripciones/canchas/lista_canchas_entrenador.html', {
        'canchas':     canchas,
        'total':       paginator.count,
        'nombre':      nombre,
        'localidad':   localidad,
        'disciplina':  disciplina,
        'disciplinas': Cancha.TipoDisciplina.choices,
        'canchas_json': json.dumps(canchas_json),
    })
