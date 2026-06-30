from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Q
from django.utils.html import escape
from datetime import date
from copy import copy
from io import BytesIO
from openpyxl import Workbook

from .models import Torneo, InscripcionTorneo, Partido, EstadisticaJugador
from .forms import TorneoForm, PartidoForm, EstadisticaForm
from inscripciones.models import Cancha, Equipo
from inscripciones.seleccion_equipo import equipo_activo
from inscripciones.seleccion_equipo import equipos_del_entrenador
from accounts.models import Jugador


def _emails_jugadores_equipo(equipo):
    if not equipo:
        return []
    return list(
        Jugador.objects.filter(equipo=equipo, is_active=True)
        .exclude(_email='')
        .values_list('_email', flat=True)
    )


def _enviar_correo_partido(partido, creado=True):
    equipos = [partido.equipo_local, partido.equipo_visita]
    destinatarios = []
    for equipo in equipos:
        destinatarios.extend(_emails_jugadores_equipo(equipo))

    destinatarios = sorted(set(destinatarios))
    if not destinatarios:
        return 0

    asunto = 'Partido programado en F.A.S' if creado else 'Partido actualizado en F.A.S'
    local = partido.equipo_local.nombre if partido.equipo_local else 'Por definir'
    visita = partido.equipo_visita.nombre if partido.equipo_visita else 'Por definir'
    fecha = partido.fecha.strftime('%d/%m/%Y %H:%M')
    ubicacion = partido.ubicacion or partido.torneo.ubicacion or 'Por confirmar'
    texto = (
        f'Hola,\n\n'
        f'Tu equipo tiene un partido en {partido.torneo.nombre}.\n\n'
        f'Partido: {local} vs {visita}\n'
        f'Fecha: {fecha}\n'
        f'Lugar: {ubicacion}\n'
        f'Estado: {partido.get_estado_display()}\n\n'
        f'Revisa F.A.S para mas detalles.\n'
    )
    html = f"""
    <div style="font-family:Arial,sans-serif;background:#f4f7fb;padding:24px;color:#172033;">
      <div style="max-width:640px;margin:auto;background:#fff;border-radius:18px;overflow:hidden;border:1px solid #e5e7eb;">
        <div style="background:#0d47a1;color:#fff;padding:20px 24px;border-bottom:4px solid #ffb300;">
          <p style="margin:0 0 6px;font-size:12px;text-transform:uppercase;letter-spacing:.08em;color:#fff4cc;">Football Association System</p>
          <h2 style="margin:0;font-size:24px;">{escape(asunto)}</h2>
        </div>
        <div style="padding:24px;">
          <p>Tu equipo tiene un partido en <strong>{escape(partido.torneo.nombre)}</strong>.</p>
          <div style="background:#edf5ff;border-left:5px solid #ffb300;border-radius:12px;padding:16px;">
            <p style="font-size:18px;margin:0 0 12px;"><strong>{escape(local)}</strong> vs <strong>{escape(visita)}</strong></p>
            <p><strong>Fecha:</strong> {escape(fecha)}</p>
            <p><strong>Lugar:</strong> {escape(ubicacion)}</p>
            <p><strong>Estado:</strong> {escape(partido.get_estado_display())}</p>
          </div>
          <p style="color:#667085;">Revisa F.A.S para confirmar cualquier cambio antes de salir hacia la cancha.</p>
        </div>
      </div>
    </div>
    """
    msg = EmailMultiAlternatives(asunto, texto, settings.DEFAULT_FROM_EMAIL, destinatarios)
    msg.attach_alternative(html, 'text/html')
    msg.send(fail_silently=True)
    return len(destinatarios)


def _pdf_escape(value):
    return str(value).replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _generar_pdf_reporte(equipo, reporte, filtros=None):
    filtros = filtros or {}
    filtros_aplicados = ', '.join(
        f'{clave}: {valor}' for clave, valor in filtros.items() if valor
    ) or 'Sin filtros'
    lines = [
        'Reporte de jugadores F.A.S',
        f'Equipo: {equipo.nombre}',
        f'Fecha: {date.today().strftime("%d/%m/%Y")}',
        f'Filtros: {filtros_aplicados}',
        f'Total jugadores: {len(reporte)}',
        '',
        'Jugador | Posicion | PJ | Goles | Asistencias | Minutos | Disciplina | Puntaje',
    ]
    for item in reporte:
        jugador = item['jugador']
        lines.append(
            f"#{jugador.dorsal} {jugador.nombres} {jugador.apellidos} | "
            f"{jugador.posicion.title()} | {item['partidos']} | {item['goles']} | "
            f"{item['asistencias']} | {item['minutos']} | {item['disciplina']} | {item['puntaje']}"
        )

    stream = ['BT', '/F1 11 Tf', '50 790 Td', '14 TL']
    for index, line in enumerate(lines[:48]):
        if index:
            stream.append('T*')
        stream.append(f'({_pdf_escape(line)}) Tj')
    stream.append('ET')
    content = '\n'.join(stream).encode('latin-1', errors='replace')

    objects = []
    objects.append(b'<< /Type /Catalog /Pages 2 0 R >>')
    objects.append(b'<< /Type /Pages /Kids [3 0 R] /Count 1 >>')
    objects.append(b'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>')
    objects.append(b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>')
    objects.append(b'<< /Length ' + str(len(content)).encode('ascii') + b' >>\nstream\n' + content + b'\nendstream')

    pdf = BytesIO()
    pdf.write(b'%PDF-1.4\n')
    offsets = [0]
    for number, obj in enumerate(objects, start=1):
        offsets.append(pdf.tell())
        pdf.write(f'{number} 0 obj\n'.encode('ascii'))
        pdf.write(obj)
        pdf.write(b'\nendobj\n')
    xref = pdf.tell()
    pdf.write(f'xref\n0 {len(objects) + 1}\n'.encode('ascii'))
    pdf.write(b'0000000000 65535 f \n')
    for offset in offsets[1:]:
        pdf.write(f'{offset:010d} 00000 n \n'.encode('ascii'))
    pdf.write(f'trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF'.encode('ascii'))
    return pdf.getvalue()


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def _calcular_tabla(torneo):
    equipos_ids = torneo.inscripciones.filter(
        estado='ACTIVA'
    ).values_list('equipo_id', flat=True)
    equipos = Equipo.objects.filter(id__in=equipos_ids)

    tabla = []
    for equipo in equipos:
        pl = Partido.objects.filter(
            torneo=torneo, equipo_local=equipo,
            estado='FINALIZADO', fase='GRUPOS'
        )
        pv = Partido.objects.filter(
            torneo=torneo, equipo_visita=equipo,
            estado='FINALIZADO', fase='GRUPOS'
        )
        pj = pl.count() + pv.count()
        gf = (sum(p.goles_local  for p in pl) +
              sum(p.goles_visita for p in pv))
        gc = (sum(p.goles_visita for p in pl) +
              sum(p.goles_local  for p in pv))
        pg = pe = pp = 0
        for p in pl:
            if   p.goles_local > p.goles_visita:  pg += 1
            elif p.goles_local < p.goles_visita:  pp += 1
            else:                                  pe += 1
        for p in pv:
            if   p.goles_visita > p.goles_local:  pg += 1
            elif p.goles_visita < p.goles_local:  pp += 1
            else:                                  pe += 1
        tabla.append({
            'equipo': equipo,
            'pj': pj, 'pg': pg, 'pe': pe, 'pp': pp,
            'gf': gf, 'gc': gc, 'dg': gf - gc,
            'pts': pg * 3 + pe,
        })
    tabla.sort(key=lambda x: (-x['pts'], -x['dg'], -x['gf']))
    return tabla


def _fixture_todos_contra_todos(equipos):
    """
    Algoritmo de rotación de círculo. Garantiza que cada par de equipos
    se enfrenta exactamente una vez. Devuelve lista de (local, visita, jornada).
    """
    equipos = list(equipos)
    n = len(equipos)
    if n < 2:
        return []
    if n % 2 != 0:
        equipos.append(None)
        n += 1

    mitad    = n // 2
    fijo     = equipos[0]
    rotantes = equipos[1:]
    partidos = []

    for jornada in range(1, n):
        for i in range(mitad):
            local  = fijo if i == 0 else rotantes[i - 1]
            visita = rotantes[n - 2 - i]
            if local is not None and visita is not None:
                if jornada % 2 == 0:
                    local, visita = visita, local
                partidos.append((local, visita, jornada))
        rotantes = [rotantes[-1]] + rotantes[:-1]

    return partidos


def _clasificados_para_siguiente(torneo, n):
    """Retorna los primeros n equipos de la tabla de grupos."""
    tabla = _calcular_tabla(torneo)
    return [fila['equipo'] for fila in tabla[:n]]


def _grupos_completos(torneo):
    """True si todos los partidos de grupos están finalizados o suspendidos."""
    grupos = torneo.partidos.filter(fase='GRUPOS')
    if not grupos.exists():
        return False
    return not grupos.exclude(
        estado__in=['FINALIZADO', 'SUSPENDIDO']
    ).exists()


def _fase_completa(torneo, fase):
    partidos = torneo.partidos.filter(fase=fase)
    if not partidos.exists():
        return False
    return not partidos.exclude(
        estado__in=['FINALIZADO', 'SUSPENDIDO']
    ).exists()


def _sincronizar_torneos(qs):
    torneos = list(qs)
    for torneo in torneos:
        torneo.actualizar_estado()
    return torneos


def _construir_indicador_rendimiento(jugador_obj):
    rendimiento_vacio = {
        'nivel': 'Sin datos',
        'puntaje': 0,
        'partidos_evaluados': 0,
        'descripcion': 'Aún no hay suficientes estadísticas para calcular el nivel.',
        'metricas': [
            {'label': 'Definición', 'value': 0},
            {'label': 'Creación', 'value': 0},
            {'label': 'Ritmo', 'value': 0},
            {'label': 'Disciplina', 'value': 0},
        ],
    }
    if not jugador_obj:
        return rendimiento_vacio

    agg = EstadisticaJugador.objects.filter(jugador=jugador_obj).aggregate(
        total_goles=Sum('goles'),
        total_asistencias=Sum('asistencias'),
        total_amarillas=Sum('tarjetas_amarillas'),
        total_rojas=Sum('tarjetas_rojas'),
        total_minutos=Sum('minutos_jugados'),
    )

    partidos = EstadisticaJugador.objects.filter(jugador=jugador_obj).count()
    if partidos == 0:
        return rendimiento_vacio

    total_goles = agg['total_goles'] or 0
    total_asistencias = agg['total_asistencias'] or 0
    total_amarillas = agg['total_amarillas'] or 0
    total_rojas = agg['total_rojas'] or 0
    total_minutos = agg['total_minutos'] or 0

    promedio_goles = total_goles / partidos
    promedio_asistencias = total_asistencias / partidos
    promedio_minutos = total_minutos / partidos
    promedio_amarillas = total_amarillas / partidos
    promedio_rojas = total_rojas / partidos

    definicion = min(100, round(promedio_goles * 100))
    creacion = min(100, round(promedio_asistencias * 100))
    ritmo = min(100, round((promedio_minutos / 90) * 100))
    disciplina = max(0, min(100, round(100 - ((promedio_amarillas * 18) + (promedio_rojas * 45)))))

    puntaje = round((definicion * 0.35) + (creacion * 0.25) + (ritmo * 0.20) + (disciplina * 0.20))

    if puntaje < 25:
        nivel = 'Bajo'
        descripcion = 'Tu impacto aún está por debajo del esperado. Sigue sumando minutos y participación.'
    elif puntaje < 50:
        nivel = 'Regular'
        descripcion = 'Vas construyendo un rendimiento estable, pero todavía hay margen para crecer.'
    elif puntaje < 75:
        nivel = 'Promedio'
        descripcion = 'Estás compitiendo en un buen nivel y aportando de forma constante al equipo.'
    else:
        nivel = 'Alto'
        descripcion = 'Tu rendimiento es sobresaliente y estás marcando diferencia en el juego.'

    return {
        'nivel': nivel,
        'puntaje': puntaje,
        'partidos_evaluados': partidos,
        'descripcion': descripcion,
        'metricas': [
            {'label': 'Definición', 'value': definicion},
            {'label': 'Creación', 'value': creacion},
            {'label': 'Ritmo', 'value': ritmo},
            {'label': 'Disciplina', 'value': disciplina},
        ],
    }


def _serializar_resumen_jugador(jugador_obj, equipo):
    torneos_data = []
    if equipo:
        inscripciones = InscripcionTorneo.objects.filter(
            equipo=equipo
        ).select_related('torneo').order_by('-torneo__fecha_inicio')

        for insc in inscripciones:
            torneo = insc.torneo
            torneo.actualizar_estado()
            estado_label = 'Registrado' if torneo.estado == Torneo.Estado.PROXIMO else torneo.get_estado_display()

            partidos_equipo = Partido.objects.filter(
                torneo=torneo
            ).filter(Q(equipo_local=equipo) | Q(equipo_visita=equipo))

            stats = EstadisticaJugador.objects.filter(
                jugador=jugador_obj,
                partido__torneo=torneo,
            ).aggregate(
                total_goles=Sum('goles'),
                total_asistencias=Sum('asistencias'),
                total_amarillas=Sum('tarjetas_amarillas'),
                total_rojas=Sum('tarjetas_rojas'),
                total_minutos=Sum('minutos_jugados'),
            )

            torneos_data.append({
                'id': torneo.id,
                'nombre': torneo.nombre,
                'categoria': torneo.get_categoria_display(),
                'estado': torneo.estado,
                'estado_label': estado_label,
                'fecha_inicio': torneo.fecha_inicio.strftime('%d/%m/%Y'),
                'fecha_fin': torneo.fecha_fin.strftime('%d/%m/%Y'),
                'ubicacion': torneo.ubicacion,
                'partidos_jugados': partidos_equipo.filter(estado='FINALIZADO').count(),
                'partidos_restantes': partidos_equipo.exclude(
                    estado__in=['FINALIZADO', 'SUSPENDIDO']
                ).count(),
                'goles': stats['total_goles'] or 0,
                'asistencias': stats['total_asistencias'] or 0,
                'amarillas': stats['total_amarillas'] or 0,
                'rojas': stats['total_rojas'] or 0,
                'minutos': stats['total_minutos'] or 0,
            })

    totales = {
        'total_goles': 0,
        'total_asistencias': 0,
        'total_rojas': 0,
        'total_minutos': 0,
    }
    if jugador_obj:
        agg = EstadisticaJugador.objects.filter(jugador=jugador_obj).aggregate(
            total_goles=Sum('goles'),
            total_asistencias=Sum('asistencias'),
            total_rojas=Sum('tarjetas_rojas'),
            total_minutos=Sum('minutos_jugados'),
        )
        totales = {k: v or 0 for k, v in agg.items()}

    return {
        'equipo': equipo,
        'torneos_data': torneos_data,
        'totales': totales,
        'rendimiento': _construir_indicador_rendimiento(jugador_obj),
    }


# ─────────────────────────────────────────────
#  ADMIN — TORNEOS
# ─────────────────────────────────────────────

@login_required
def admin_lista_torneos(request):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    for t in Torneo.objects.exclude(estado__in=['CANCELADO', 'FINALIZADO']):
        t.actualizar_estado()

    nombre = request.GET.get('nombre', '').strip()
    estado = request.GET.get('estado', '').strip()
    motivo = request.GET.get('motivo', '').strip()
    detalle = request.GET.get('detalle', '').strip()
    torneos = Torneo.objects.all()
    if nombre:
        torneos = torneos.filter(nombre__icontains=nombre)
    if estado:
        torneos = torneos.filter(estado=estado)
    if motivo:
        torneos = torneos.filter(motivo_cancelacion=motivo)
    if detalle:
        torneos = torneos.filter(
            Q(motivo_cancelacion_detalle__icontains=detalle) |
            Q(descripcion__icontains=detalle)
        )

    return render(request, 'torneos/admin/lista_torneos.html', {
        'torneos':         torneos,
        'request_nombre':  nombre,
        'request_estado':  estado,
        'request_motivo':  motivo,
        'request_detalle': detalle,
        'estados':         Torneo.Estado.choices,
        'motivos_cancelacion': Torneo.MotivoCancelacion.choices,
    })


@login_required
def admin_crear_torneo(request):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    form = TorneoForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Torneo creado exitosamente.')
        return redirect('torneos:admin_lista_torneos')

    return render(request, 'torneos/admin/crear_torneo.html', {'form': form})


@login_required
def admin_canchas_disponibles(request):
    if request.user.rol != 'ADMIN':
        return JsonResponse({'canchas': []}, status=403)

    canchas = Cancha.objects.filter(
        _disponibilidad=Cancha.Disponibilidad.DISPONIBLE
    ).order_by('_nombre_escenario')

    return JsonResponse({
        'canchas': [
            {
                'id': cancha.id,
                'nombre': cancha.nombre_escenario,
                'direccion': cancha.direccion_exacta,
                'localidad': cancha.localidad,
                'barrio': cancha.barrio,
                'lat': cancha.latitud,
                'lng': cancha.longitud,
                'ubicacion': f'{cancha.nombre_escenario} - {cancha.direccion_exacta}',
            }
            for cancha in canchas
        ]
    })


@login_required
def admin_editar_torneo(request, torneo_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    torneo = get_object_or_404(Torneo, id=torneo_id)
    form   = TorneoForm(request.POST or None, instance=torneo)

    if request.method == 'POST' and form.is_valid():
        torneo = form.save()
        torneo.actualizar_estado()
        messages.success(request, 'Torneo actualizado exitosamente.')
        return redirect('torneos:admin_lista_torneos')

    return render(request, 'torneos/admin/editar_torneo.html', {
        'form': form, 'torneo': torneo, 'editando': True
    })


@login_required
def admin_eliminar_torneo(request, torneo_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    torneo = get_object_or_404(Torneo, id=torneo_id)
    if request.method == 'POST':
        motivo = request.POST.get('motivo_cancelacion', '').strip()
        detalle = request.POST.get('motivo_cancelacion_detalle', '').strip()
        motivos_validos = {choice[0] for choice in Torneo.MotivoCancelacion.choices}

        if motivo not in motivos_validos:
            messages.error(request, 'Selecciona un motivo para cancelar el torneo.')
            return redirect('torneos:admin_lista_torneos')

        torneo.estado = Torneo.Estado.CANCELADO
        torneo.motivo_cancelacion = motivo
        torneo.motivo_cancelacion_detalle = detalle
        torneo.save(update_fields=['estado', 'motivo_cancelacion', 'motivo_cancelacion_detalle'])
        messages.success(request, 'Torneo cancelado y conservado en el historial.')
    return redirect('torneos:admin_lista_torneos')


@login_required
def admin_detalle_torneo(request, torneo_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    torneo = get_object_or_404(Torneo, id=torneo_id)
    torneo.actualizar_estado()

    partidos_grupos    = torneo.partidos.filter(fase='GRUPOS').order_by('jornada', 'fecha')
    partidos_cuartos   = torneo.partidos.filter(fase='CUARTOS').order_by('fecha')
    partidos_semi      = torneo.partidos.filter(fase='SEMIFINAL').order_by('fecha')
    partidos_tercero   = torneo.partidos.filter(fase='TERCER_PUES').order_by('fecha')
    partidos_final     = torneo.partidos.filter(fase='FINAL').order_by('fecha')
    inscritos          = torneo.inscripciones.filter(estado='ACTIVA').select_related('equipo')
    tabla              = _calcular_tabla(torneo)

    # Qué botones mostrar
    puede_fixture_grupos   = (
        not torneo.partidos.filter(fase='GRUPOS').exists()
        and inscritos.count() >= 2
    )
    puede_avanzar_cuartos  = (
        torneo.formato == 'GRUPOS_CUARTOS'
        and _grupos_completos(torneo)
        and not torneo.partidos.filter(fase='CUARTOS').exists()
        and len(tabla) >= 8
    )
    puede_avanzar_semi     = (
        torneo.formato in ['GRUPOS_SEMI', 'GRUPOS_CUARTOS']
        and not torneo.partidos.filter(fase='SEMIFINAL').exists()
        and (
            (_fase_completa(torneo, 'CUARTOS') and torneo.partidos.filter(fase='CUARTOS').exists())
            or
            (_grupos_completos(torneo) and torneo.formato == 'GRUPOS_SEMI' and not torneo.partidos.filter(fase='CUARTOS').exists() and len(tabla) >= 4)
        )
    )
    puede_avanzar_final    = (
        torneo.formato in ['GRUPOS_FINAL', 'GRUPOS_SEMI', 'GRUPOS_CUARTOS']
        and not torneo.partidos.filter(fase='FINAL').exists()
        and (
            (_fase_completa(torneo, 'SEMIFINAL') and torneo.partidos.filter(fase='SEMIFINAL').exists())
            or
            (_grupos_completos(torneo) and torneo.formato == 'GRUPOS_FINAL' and len(tabla) >= 2)
        )
    )

    partidos_finalizados = torneo.partidos.filter(estado='FINALIZADO').count()

    return render(request, 'torneos/admin/detalle_torneo.html', {
        'torneo':                torneo,
        'inscritos':             inscritos,
        'tabla':                 tabla,
        'partidos_grupos':       partidos_grupos,
        'partidos_cuartos':      partidos_cuartos,
        'partidos_semi':         partidos_semi,
        'partidos_tercero':      partidos_tercero,
        'partidos_final':        partidos_final,
        'puede_fixture_grupos':  puede_fixture_grupos,
        'puede_avanzar_cuartos': puede_avanzar_cuartos,
        'puede_avanzar_semi':    puede_avanzar_semi,
        'puede_avanzar_final':   puede_avanzar_final,
        'partidos_finalizados':  partidos_finalizados,
    })


# ─────────────────────────────────────────────
#  ADMIN — PARTIDOS
# ─────────────────────────────────────────────

@login_required
def admin_crear_partido(request, torneo_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    torneo = get_object_or_404(Torneo, id=torneo_id)
    form   = PartidoForm(request.POST or None, torneo=torneo)

    if request.method == 'POST' and form.is_valid():
        partido = form.save(commit=False)
        partido.torneo = torneo
        partido.save()
        enviados = _enviar_correo_partido(partido, creado=True)
        if enviados:
            messages.success(request, f'Partido creado. Se notifico a {enviados} jugador(es).')
        else:
            messages.success(request, 'Partido creado. No habia correos de jugadores para notificar.')
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    return render(request, 'torneos/admin/crear_partido.html', {
        'form': form, 'torneo': torneo
    })


@login_required
def admin_editar_partido(request, partido_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    partido = get_object_or_404(Partido, id=partido_id)
    form    = PartidoForm(
        request.POST or None,
        instance=partido,
        torneo=partido.torneo
    )

    if request.method == 'POST' and form.is_valid():
        partido = form.save()
        enviados = _enviar_correo_partido(partido, creado=False)
        if enviados:
            messages.success(request, f'Partido actualizado. Se notifico a {enviados} jugador(es).')
        else:
            messages.success(request, 'Partido actualizado. No habia correos de jugadores para notificar.')
        return redirect('torneos:admin_detalle_torneo', torneo_id=partido.torneo.id)

    return render(request, 'torneos/admin/crear_partido.html', {
        'form': form, 'torneo': partido.torneo, 'editando': True
    })


@login_required
def admin_desinscribir_equipo(request, inscripcion_id):
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    inscripcion = get_object_or_404(InscripcionTorneo, id=inscripcion_id)
    torneo_id   = inscripcion.torneo.id
    if request.method == 'POST':
        inscripcion.estado = InscripcionTorneo.Estado.CANCELADA
        inscripcion.save()
        messages.success(request, 'Equipo desinscrito.')
    return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)


# ─────────────────────────────────────────────
#  ADMIN — GENERACIÓN DE FASES
# ─────────────────────────────────────────────

@login_required
def admin_generar_fixture(request, torneo_id):
    """Genera la fase de grupos (todos contra todos)."""
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    torneo = get_object_or_404(Torneo, id=torneo_id)

    if request.method != 'POST':
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    if torneo.partidos.filter(fase='GRUPOS').exists():
        messages.error(request, 'Ya existe una fase de grupos para este torneo.')
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    equipos = list(
        Equipo.objects.filter(
            id__in=torneo.inscripciones.filter(
                estado='ACTIVA'
            ).values_list('equipo_id', flat=True)
        )
    )

    if len(equipos) < 2:
        messages.error(request, 'Se necesitan al menos 2 equipos inscritos.')
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    pares_vistos = set()
    partidos_a_crear = _fixture_todos_contra_todos(equipos)

    creados = 0
    for local, visita, jornada in partidos_a_crear:
        par = tuple(sorted([local.id, visita.id]))
        if par in pares_vistos:
            continue
        pares_vistos.add(par)
        Partido.objects.create(
            torneo=torneo,
            equipo_local=local,
            equipo_visita=visita,
            fase=Partido.Fase.GRUPOS,
            jornada=jornada,
            fecha=torneo.fecha_inicio,
            estado=Partido.Estado.PROGRAMADO,
        )
        creados += 1

    torneo.fase_actual = 'GRUPOS'
    torneo.save(update_fields=['fase_actual'])

    n_jornadas = max(j for _, _, j in partidos_a_crear) if partidos_a_crear else 0
    messages.success(
        request,
        f'Fase de grupos generada: {creados} partido(s) en {n_jornadas} jornada(s). '
        f'Asigna las fechas exactas editando cada partido.'
    )
    return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)


@login_required
def admin_generar_cuartos(request, torneo_id):
    """Genera cuartos de final con los 8 primeros de la tabla."""
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    torneo = get_object_or_404(Torneo, id=torneo_id)

    if request.method != 'POST':
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    if not _grupos_completos(torneo):
        messages.error(request, 'Todos los partidos de grupos deben estar finalizados.')
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    if torneo.partidos.filter(fase='CUARTOS').exists():
        messages.error(request, 'Los cuartos de final ya fueron generados.')
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    clasificados = _clasificados_para_siguiente(torneo, 8)
    if len(clasificados) < 8:
        messages.error(request, f'Se necesitan 8 equipos clasificados. Solo hay {len(clasificados)}.')
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    # 1° vs 8°, 2° vs 7°, 3° vs 6°, 4° vs 5°
    enfrentamientos = [
        (clasificados[0], clasificados[7]),
        (clasificados[1], clasificados[6]),
        (clasificados[2], clasificados[5]),
        (clasificados[3], clasificados[4]),
    ]
    for local, visita in enfrentamientos:
        Partido.objects.create(
            torneo=torneo,
            equipo_local=local,
            equipo_visita=visita,
            fase=Partido.Fase.CUARTOS,
            jornada=1,
            fecha=torneo.fecha_inicio,
            estado=Partido.Estado.PROGRAMADO,
        )

    torneo.fase_actual = 'CUARTOS'
    torneo.save(update_fields=['fase_actual'])

    messages.success(request, 'Cuartos de final generados (4 partidos). Asigna fechas y registra resultados.')
    return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)


@login_required
def admin_generar_semifinales(request, torneo_id):
    """Genera semifinales. Toma los 4 primeros o los ganadores de cuartos."""
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    torneo = get_object_or_404(Torneo, id=torneo_id)

    if request.method != 'POST':
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    if torneo.partidos.filter(fase='SEMIFINAL').exists():
        messages.error(request, 'Las semifinales ya fueron generadas.')
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    # Determinar los 4 semifinalistas
    cuartos = torneo.partidos.filter(fase='CUARTOS', estado='FINALIZADO')
    if cuartos.exists():
        if not _fase_completa(torneo, 'CUARTOS'):
            messages.error(request, 'Todos los cuartos de final deben estar finalizados.')
            return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)
        # Ganadores de cuartos en orden
        semifinalistas = []
        for p in torneo.partidos.filter(fase='CUARTOS').order_by('id'):
            if p.goles_local > p.goles_visita:
                semifinalistas.append(p.equipo_local)
            elif p.goles_visita > p.goles_local:
                semifinalistas.append(p.equipo_visita)
            else:
                messages.error(request, f'El partido {p} terminó empatado. Define un ganador antes de avanzar.')
                return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)
    else:
        if not _grupos_completos(torneo):
            messages.error(request, 'Todos los partidos de grupos deben estar finalizados.')
            return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)
        semifinalistas = _clasificados_para_siguiente(torneo, 4)
        if len(semifinalistas) < 4:
            messages.error(request, f'Se necesitan 4 equipos. Solo hay {len(semifinalistas)}.')
            return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    # Semi 1: [0] vs [3], Semi 2: [1] vs [2]
    Partido.objects.create(
        torneo=torneo, equipo_local=semifinalistas[0], equipo_visita=semifinalistas[3],
        fase=Partido.Fase.SEMIFINAL, jornada=1, fecha=torneo.fecha_inicio,
        estado=Partido.Estado.PROGRAMADO,
    )
    Partido.objects.create(
        torneo=torneo, equipo_local=semifinalistas[1], equipo_visita=semifinalistas[2],
        fase=Partido.Fase.SEMIFINAL, jornada=1, fecha=torneo.fecha_inicio,
        estado=Partido.Estado.PROGRAMADO,
    )

    torneo.fase_actual = 'SEMIFINAL'
    torneo.save(update_fields=['fase_actual'])

    messages.success(request, 'Semifinales generadas (2 partidos). Registra los resultados para avanzar a la final.')
    return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)


@login_required
def admin_generar_final(request, torneo_id):
    """Genera la final y el partido por el tercer puesto."""
    if request.user.rol != 'ADMIN':
        return redirect('dashboard_entrenador')

    torneo = get_object_or_404(Torneo, id=torneo_id)

    if request.method != 'POST':
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    if torneo.partidos.filter(fase='FINAL').exists():
        messages.error(request, 'La final ya fue generada.')
        return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

    semis = torneo.partidos.filter(fase='SEMIFINAL')

    if semis.exists():
        if not _fase_completa(torneo, 'SEMIFINAL'):
            messages.error(request, 'Ambas semifinales deben estar finalizadas.')
            return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

        finalistas   = []
        perdedores   = []
        for p in semis.order_by('id'):
            if p.goles_local > p.goles_visita:
                finalistas.append(p.equipo_local)
                perdedores.append(p.equipo_visita)
            elif p.goles_visita > p.goles_local:
                finalistas.append(p.equipo_visita)
                perdedores.append(p.equipo_local)
            else:
                messages.error(request, f'La semifinal {p} terminó empatada. Define un ganador.')
                return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

        if len(finalistas) < 2:
            messages.error(request, 'No se pudieron determinar los finalistas.')
            return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)

        # Tercer puesto
        Partido.objects.create(
            torneo=torneo, equipo_local=perdedores[0], equipo_visita=perdedores[1],
            fase=Partido.Fase.TERCER_PUES, jornada=1, fecha=torneo.fecha_fin,
            estado=Partido.Estado.PROGRAMADO,
        )

    else:
        # Formato GRUPOS_FINAL: los 2 primeros van directo a la final
        if not _grupos_completos(torneo):
            messages.error(request, 'Todos los partidos de grupos deben estar finalizados.')
            return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)
        clasificados = _clasificados_para_siguiente(torneo, 2)
        if len(clasificados) < 2:
            messages.error(request, 'Se necesitan al menos 2 equipos clasificados.')
            return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)
        finalistas = clasificados

    Partido.objects.create(
        torneo=torneo, equipo_local=finalistas[0], equipo_visita=finalistas[1],
        fase=Partido.Fase.FINAL, jornada=1, fecha=torneo.fecha_fin,
        estado=Partido.Estado.PROGRAMADO,
    )

    torneo.fase_actual = 'FINAL'
    torneo.save(update_fields=['fase_actual'])

    messages.success(request, 'Final generada. ¡Registra el resultado para coronar al campeón!')
    return redirect('torneos:admin_detalle_torneo', torneo_id=torneo_id)


# ─────────────────────────────────────────────
#  ENTRENADOR
# ─────────────────────────────────────────────

@login_required
def entrenador_lista_torneos(request):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    equipo = equipo_activo(request)

    mis_inscripciones = []
    ids_inscritos     = []
    if equipo:
        mis_inscripciones_qs = InscripcionTorneo.objects.filter(
            equipo=equipo, estado='ACTIVA'
        ).select_related('torneo')
        ids_inscritos = list(mis_inscripciones_qs.values_list('torneo_id', flat=True))
        mis_inscripciones = list(mis_inscripciones_qs)
        for inscripcion in mis_inscripciones:
            inscripcion.torneo.actualizar_estado()

    torneos_qs = Torneo.objects.filter(
        cupo_maximo__gt=0,
    ).exclude(id__in=ids_inscritos)

    if equipo:
        torneos_qs = torneos_qs.filter(categoria=equipo.categoria)

    torneos_disponibles = [
        t for t in _sincronizar_torneos(torneos_qs)
        if t.puede_inscribirse
    ]

    return render(request, 'torneos/entrenador/lista_torneos.html', {
        'torneos_disponibles': torneos_disponibles,
        'mis_inscripciones':   mis_inscripciones,
        'equipo':              equipo,
        'equipos':             equipos_del_entrenador(request.user),
    })


@login_required
def entrenador_inscribir(request, torneo_id):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    torneo = get_object_or_404(Torneo, id=torneo_id)
    equipo = equipo_activo(request)

    if not equipo:
        messages.error(request, 'No tienes un equipo registrado.')
        return redirect('torneos:entrenador_lista_torneos')

    if equipo.estado != 'APROBADO':
        messages.error(request, 'Tu equipo debe estar aprobado para inscribirse.')
        return redirect('torneos:entrenador_lista_torneos')

    if equipo.categoria != torneo.categoria:
        messages.error(
            request,
            f'Tu equipo es categoría {equipo.categoria_display} y este '
            f'torneo es para {torneo.get_categoria_display()}.'
        )
        return redirect('torneos:entrenador_lista_torneos')

    if not torneo.puede_inscribirse:
        messages.error(request, 'Este torneo no acepta inscripciones.')
        return redirect('torneos:entrenador_lista_torneos')

    inscripcion_existente = InscripcionTorneo.objects.filter(
        torneo=torneo, equipo=equipo
    ).first()

    if inscripcion_existente:
        if inscripcion_existente.estado == InscripcionTorneo.Estado.ACTIVA:
            messages.error(request, 'Tu equipo ya está inscrito en este torneo.')
            return redirect('torneos:entrenador_lista_torneos')

        if request.method == 'POST':
            inscripcion_existente.estado = InscripcionTorneo.Estado.ACTIVA
            inscripcion_existente.save()
            messages.success(request, f'¡Tu equipo fue inscrito en {torneo.nombre}!')
        return redirect('torneos:entrenador_lista_torneos')

    if request.method == 'POST':
        InscripcionTorneo.objects.create(torneo=torneo, equipo=equipo)
        messages.success(request, f'¡Tu equipo fue inscrito en {torneo.nombre}!')

    return redirect('torneos:entrenador_lista_torneos')


@login_required
def entrenador_cancelar_inscripcion(request, inscripcion_id):
    if request.user.rol != 'ADMIN' and request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    inscripcion = get_object_or_404(InscripcionTorneo, id=inscripcion_id)
    equipo      = equipo_activo(request)

    if request.user.rol == 'ENTRENADOR' and inscripcion.equipo != equipo:
        messages.error(request, 'No tienes permiso para esta acción.')
        return redirect('torneos:entrenador_lista_torneos')

    if inscripcion.torneo.estado != Torneo.Estado.PROXIMO:
        messages.error(request, 'No puedes cancelar un torneo que ya inició.')
        return redirect('torneos:entrenador_lista_torneos')

    if request.method == 'POST':
        inscripcion.estado = InscripcionTorneo.Estado.CANCELADA
        inscripcion.save()
        messages.success(request, 'Inscripción cancelada.')

    return redirect('torneos:entrenador_lista_torneos')


@login_required
def entrenador_confirmar_cancelar(request, inscripcion_id):
    if request.user.rol != 'ADMIN' and request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    inscripcion = get_object_or_404(
        InscripcionTorneo.objects.select_related('torneo', 'equipo'),
        id=inscripcion_id
    )
    equipo = equipo_activo(request)

    if request.user.rol == 'ENTRENADOR' and inscripcion.equipo != equipo:
        messages.error(request, 'No tienes permiso para esta acción.')
        return redirect('torneos:entrenador_lista_torneos')

    if inscripcion.estado != InscripcionTorneo.Estado.ACTIVA:
        messages.error(request, 'Esta inscripción ya no está activa.')
        return redirect('torneos:entrenador_lista_torneos')

    if inscripcion.torneo.estado != Torneo.Estado.PROXIMO:
        messages.error(request, 'No puedes cancelar un torneo que ya inició.')
        return redirect('torneos:entrenador_lista_torneos')

    if request.method == 'POST':
        inscripcion.estado = InscripcionTorneo.Estado.CANCELADA
        inscripcion.save()
        messages.success(request, f'Inscripción cancelada en {inscripcion.torneo.nombre}.')
        return redirect('torneos:entrenador_lista_torneos')

    return render(request, 'torneos/entrenador/confirmar_cancelar.html', {
        'inscripcion': inscripcion,
        'torneo':      inscripcion.torneo,
        'equipo':      inscripcion.equipo,
    })


@login_required
def entrenador_mis_partidos(request, torneo_id):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    torneo = get_object_or_404(Torneo, id=torneo_id)
    torneo.actualizar_estado()
    equipo = equipo_activo(request)

    partidos = Partido.objects.filter(
        torneo=torneo
    ).filter(
        Q(equipo_local=equipo) | Q(equipo_visita=equipo)
    ).order_by('fase', 'jornada', 'fecha')

    pj = partidos.filter(estado='FINALIZADO').count()
    pr = partidos.exclude(estado__in=['FINALIZADO', 'SUSPENDIDO']).count()

    return render(request, 'torneos/entrenador/mis_partidos.html', {
        'torneo':             torneo,
        'partidos':           partidos,
        'equipo':             equipo,
        'partidos_jugados':   pj,
        'partidos_restantes': pr,
        'tabla':              _calcular_tabla(torneo),
    })


@login_required
def entrenador_estadisticas_partido(request, partido_id):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    partido = get_object_or_404(
        Partido.objects.select_related(
            'torneo', 'equipo_local', 'equipo_visita'
        ),
        id=partido_id
    )
    equipo = equipo_activo(request)

    if not equipo or equipo not in [partido.equipo_local, partido.equipo_visita]:
        messages.error(request, 'No tienes permisos para gestionar este partido.')
        return redirect('torneos:entrenador_lista_torneos')

    jugadores = Jugador.objects.filter(equipo=equipo).order_by('_dorsal', '_nombres')
    estadisticas_existentes = {
        estadistica.jugador_id: estadistica
        for estadistica in EstadisticaJugador.objects.filter(
            partido=partido,
            jugador__equipo=equipo
        ).select_related('jugador')
    }

    # Goles máximos que puede registrar este equipo según el marcador
    if equipo == partido.equipo_local:
        max_goles_equipo = partido.goles_local
    else:
        max_goles_equipo = partido.goles_visita

    if request.method == 'POST':
        def entero_no_negativo(nombre):
            try:
                return max(0, int(request.POST.get(nombre, 0) or 0))
            except (TypeError, ValueError):
                return 0

        # Validar que los goles ingresados no superen el marcador
        total_goles_ingresados = sum(
            entero_no_negativo(f'goles_{jugador.id}')
            for jugador in jugadores
        )

        if total_goles_ingresados > max_goles_equipo:
            messages.error(
                request,
                f'El marcador registra {max_goles_equipo} gol(es) para tu equipo, '
                f'pero intentas registrar {total_goles_ingresados}. Ajusta los valores.'
            )
        else:
            procesados = 0
            for jugador in jugadores:
                goles = entero_no_negativo(f'goles_{jugador.id}')
                asistencias = entero_no_negativo(f'asistencias_{jugador.id}')
                amarillas = entero_no_negativo(f'amarillas_{jugador.id}')
                rojas = entero_no_negativo(f'rojas_{jugador.id}')
                minutos = entero_no_negativo(f'minutos_{jugador.id}')

                estadistica, _ = EstadisticaJugador.objects.update_or_create(
                    partido=partido,
                    jugador=jugador,
                    defaults={
                        'equipo': equipo,
                        'goles': goles,
                        'asistencias': asistencias,
                        'tarjetas_amarillas': amarillas,
                        'tarjetas_rojas': rojas,
                        'minutos_jugados': minutos,
                    }
                )
                estadisticas_existentes[jugador.id] = estadistica
                procesados += 1

            messages.success(
                request,
                f'Se guardaron las estadísticas de {procesados} jugador(es).'
            )
            return redirect('torneos:entrenador_estadisticas_partido', partido_id=partido.id)

    jugadores_con_stats = []
    for jugador in jugadores:
        estadistica = estadisticas_existentes.get(jugador.id)
        jugadores_con_stats.append({
            'jugador': jugador,
            'estadistica': estadistica,
        })

    return render(request, 'torneos/entrenador/estadisticas_partido.html', {
        'partido': partido,
        'torneo': partido.torneo,
        'equipo': equipo,
        'max_goles_equipo': max_goles_equipo,
        'jugadores_con_stats': jugadores_con_stats,
    })


@login_required
def entrenador_reporte_jugadores(request):
    if request.user.rol != 'ENTRENADOR':
        return redirect('dashboard_admin')

    equipo = equipo_activo(request)
    if not equipo:
        messages.error(request, 'Necesitas un equipo para generar reportes.')
        return redirect('dashboard_entrenador')

    busqueda = request.GET.get('q', '').strip()
    posicion = request.GET.get('posicion', '').strip()
    rendimiento = request.GET.get('rendimiento', '').strip()
    orden = request.GET.get('orden', 'puntaje').strip()
    export = request.GET.get('export', '').strip().lower()

    jugadores = Jugador.objects.filter(equipo=equipo).order_by('_dorsal', '_nombres')
    if busqueda:
        jugadores = jugadores.filter(
            Q(_nombres__icontains=busqueda) |
            Q(_apellidos__icontains=busqueda) |
            Q(_num_documento__icontains=busqueda)
        )
    if posicion:
        jugadores = jugadores.filter(_posicion=posicion)

    posiciones = list(
        Jugador.objects.filter(equipo=equipo)
        .exclude(_posicion='')
        .order_by('_posicion')
        .values_list('_posicion', flat=True)
        .distinct()
    )

    reporte = []
    for jugador in jugadores:
        agg = EstadisticaJugador.objects.filter(jugador=jugador).aggregate(
            goles=Sum('goles'),
            asistencias=Sum('asistencias'),
            amarillas=Sum('tarjetas_amarillas'),
            rojas=Sum('tarjetas_rojas'),
            minutos=Sum('minutos_jugados'),
        )
        partidos = EstadisticaJugador.objects.filter(jugador=jugador).count()
        goles = agg['goles'] or 0
        asistencias = agg['asistencias'] or 0
        amarillas = agg['amarillas'] or 0
        rojas = agg['rojas'] or 0
        minutos = agg['minutos'] or 0
        disciplina = max(0, 100 - (amarillas * 8) - (rojas * 25))
        puntaje = min(100, round((goles * 12) + (asistencias * 8) + (minutos / 12) + (disciplina * 0.25)))

        reporte.append({
            'jugador': jugador,
            'partidos': partidos,
            'goles': goles,
            'asistencias': asistencias,
            'amarillas': amarillas,
            'rojas': rojas,
            'minutos': minutos,
            'disciplina': disciplina,
            'puntaje': puntaje,
        })

    if rendimiento == 'alto':
        reporte = [item for item in reporte if item['puntaje'] >= 75]
    elif rendimiento == 'medio':
        reporte = [item for item in reporte if 40 <= item['puntaje'] < 75]
    elif rendimiento == 'bajo':
        reporte = [item for item in reporte if item['puntaje'] < 40 and item['partidos'] > 0]
    elif rendimiento == 'sin_datos':
        reporte = [item for item in reporte if item['partidos'] == 0]

    ordenadores = {
        'puntaje': lambda item: (-item['puntaje'], -item['goles'], -item['asistencias'], item['jugador'].apellidos),
        'goles': lambda item: (-item['goles'], -item['puntaje'], item['jugador'].apellidos),
        'asistencias': lambda item: (-item['asistencias'], -item['puntaje'], item['jugador'].apellidos),
        'minutos': lambda item: (-item['minutos'], -item['puntaje'], item['jugador'].apellidos),
        'dorsal': lambda item: (item['jugador'].dorsal, item['jugador'].apellidos),
    }
    reporte.sort(key=ordenadores.get(orden, ordenadores['puntaje']))
    mejores = reporte[:3]
    apoyo = sorted(reporte, key=lambda item: (item['puntaje'], item['partidos']))[:3]

    if export == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Jugadores'
        ws.append(['Reporte de jugadores F.A.S'])
        ws.append([f'Equipo: {equipo.nombre}'])
        ws.append([f'Fecha: {date.today().strftime("%d/%m/%Y")}'])
        ws.append([f'Filtros: busqueda={busqueda or "todos"}; posicion={posicion or "todas"}; rendimiento={rendimiento or "todos"}; orden={orden}'])
        ws.append([])
        ws.append(['Jugador', 'Documento', 'Dorsal', 'Posicion', 'Partidos', 'Goles', 'Asistencias', 'Minutos', 'Disciplina', 'Puntaje'])
        header_row = ws.max_row
        for item in reporte:
            jugador = item['jugador']
            ws.append([
                f'{jugador.nombres} {jugador.apellidos}',
                jugador.num_documento,
                jugador.dorsal,
                jugador.posicion.title(),
                item['partidos'],
                item['goles'],
                item['asistencias'],
                item['minutos'],
                item['disciplina'],
                item['puntaje'],
            ])
        ws.freeze_panes = f'A{header_row + 1}'
        for cell in ws[header_row]:
            font = copy(cell.font)
            fill = copy(cell.fill)
            font.bold = True
            font.color = 'FFFFFF'
            fill.fill_type = 'solid'
            fill.fgColor = '0D47A1'
            cell.font = font
            cell.fill = fill
        for column in ws.columns:
            letter = column[0].column_letter
            ws.column_dimensions[letter].width = min(32, max(12, max(len(str(cell.value or '')) for cell in column) + 2))
        output = BytesIO()
        wb.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_jugadores.xlsx"'
        return response

    if export == 'pdf':
        response = HttpResponse(_generar_pdf_reporte(equipo, reporte, {
            'busqueda': busqueda,
            'posicion': posicion,
            'rendimiento': rendimiento,
            'orden': orden,
        }), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_jugadores.pdf"'
        return response

    return render(request, 'torneos/entrenador/reporte_jugadores.html', {
        'equipo': equipo,
        'equipos': equipos_del_entrenador(request.user),
        'reporte': reporte,
        'mejores': mejores,
        'apoyo': apoyo,
        'posiciones': posiciones,
        'filtros': {
            'q': busqueda,
            'posicion': posicion,
            'rendimiento': rendimiento,
            'orden': orden,
        },
    })


# ─────────────────────────────────────────────
#  JUGADOR
# ─────────────────────────────────────────────

@login_required
def jugador_mis_torneos(request):
    if request.user.rol != 'JUGADOR':
        return redirect('dashboard_admin')

    jugador_obj = getattr(request.user, 'jugador', None)
    equipo      = jugador_obj.equipo if jugador_obj else None
    resumen = _serializar_resumen_jugador(jugador_obj, equipo)

    return render(request, 'torneos/jugador/mis_torneos.html', {
        'torneos_data': resumen['torneos_data'],
        'totales':      resumen['totales'],
        'equipo':       resumen['equipo'],
        'rendimiento':  resumen['rendimiento'],
    })


@login_required
def jugador_mis_torneos_datos(request):
    if request.user.rol != 'JUGADOR':
        return JsonResponse({'detail': 'No autorizado'}, status=403)

    jugador_obj = getattr(request.user, 'jugador', None)
    equipo = jugador_obj.equipo if jugador_obj else None
    resumen = _serializar_resumen_jugador(jugador_obj, equipo)

    return JsonResponse({
        'equipo': resumen['equipo'].nombre if resumen['equipo'] else None,
        'totales': resumen['totales'],
        'torneos': resumen['torneos_data'],
        'rendimiento': resumen['rendimiento'],
    })
